"""Metan.iQ — Generatore XLSX editabile con formule live.

Crea un workbook Excel autocalcolante: l'utente modifica solo le celle
gialle (Ore + Biomasse) e tutti gli altri valori (produzione, saving GHG,
validita') si ricalcolano automaticamente IN EXCEL grazie alle formule.

Sheets:
- "Piano mensile": tabella editabile + formule
- "Database feedstock": tabella feedstock con yield, eec, etd, esca, e_total
- "Sintesi annuale": KPI aggregati con formule cross-sheet
"""
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

from i18n_runtime import t as _t


# ============================================================
# Design tokens (mirror dell'app)
# ============================================================
NAVY      = "0F172A"
NAVY_2    = "1E293B"
AMBER     = "F59E0B"
AMBER_DK  = "B45309"
AMBER_BG  = "FEF3C7"
SLATE_50  = "F8FAFC"
SLATE_100 = "F1F5F9"
SLATE_200 = "E2E8F0"
SLATE_400 = "94A3B8"
SLATE_500 = "64748B"
SLATE_700 = "334155"
EMERALD_BG = "D1FAE5"
EMERALD_FG = "065F46"
RED_BG     = "FECACA"
RED_FG     = "991B1B"
WHITE     = "FFFFFF"


def _border_thin():
    side = Side(style="thin", color=SLATE_200)
    return Border(left=side, right=side, top=side, bottom=side)




def _style_header(c):
    c.font = Font(bold=True, color=WHITE, size=10)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True,
    )
    c.border = _border_thin()


def _style_editable(c):
    c.fill = PatternFill("solid", fgColor=AMBER_BG)
    c.border = _border_thin()
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.font = Font(color=NAVY)


def _style_readonly(c):
    c.fill = PatternFill("solid", fgColor=SLATE_50)
    c.border = _border_thin()
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.font = Font(color=SLATE_700)


def _style_total(c):
    c.font = Font(bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = _border_thin()


# ============================================================
# Public API
# ============================================================
def build_metaniq_xlsx(ctx: dict, snapshot: bool = False) -> BytesIO:
    """Costruisce il workbook XLSX completo.

    Parametri:
      ctx: dict con active_feeds, FEEDSTOCK_DB, aux_factor, ep_total,
           fossil_comparator, ghg_threshold, plant_net_smch, MONTHS,
           MONTH_HOURS, NM3_TO_MWH, end_use, APP_MODE_LABEL,
           initial_data (per editabile) o df_res (per snapshot).

      snapshot: False (default) -> file EDITABILE con formule live.
                True -> file SNAPSHOT con valori statici (no formule).
    """
    wb = Workbook()
    lang = ctx.get("lang", "it")  # lingua: 'it' o 'en'

    # Forza ricalcolo completo all'apertura del file in Excel.
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    # === Sheet 1: Database (creata per prima per nome reference) ===
    ws_db = wb.create_sheet(_t("Database feedstock", lang))
    _build_database(ws_db, ctx, lang=lang)

    # === Sheet 2: Piano (main, attiva di default) ===
    ws_piano = wb.active
    ws_piano.title = _t("Piano mensile", lang)
    _build_piano(ws_piano, ctx, ws_db.title, snapshot=snapshot, lang=lang)

    # === Sheet 3: Sintesi ===
    ws_sum = wb.create_sheet(_t("Sintesi annuale", lang))
    _build_summary(ws_sum, ctx, ws_piano.title, lang=lang)

    # === Sheet 4: Business Plan (sempre, anche per snapshot) ===
    # Pro forma 15 anni con formule live: CAPEX, OPEX, CE, cash flow, KPI.
    # Mode-aware: legge taglia + tariffa + autoconsumi dal ctx.
    # ws_bp = wb.create_sheet("Business Plan")
    # _build_business_plan(ws_bp, ctx, snapshot=snapshot, lang=lang)

    # === Sheet 5: Audit Rese BMT (override certificati vs tabella standard) ===
    # Tracciabilita' completa: per ogni biomassa attiva mostra resa standard,
    # resa usata, origine (tabella vs BMT certificato laboratorio), nome
    # certificato, laboratorio, data, riferimento campione.
    ws_bmt = wb.create_sheet(_t("Audit rese BMT", lang) if lang == "en" else "Audit rese BMT")
    _build_bmt_audit(ws_bmt, ctx, lang=lang)
    # === Sheet 5: Audit fattori emissivi (override REALI vs standard) ===
    # Tracciabilita' completa: per ogni biomassa attiva mostra fattori
    # standard, fattori usati, scostamento %, origine (relazione vs
    # tabella), metadati relazione (titolo, autore, societa', data,
    # impianto, riferimento campione, note metodologiche).
    if ctx.get("emission_audit_rows"):
        ws_ef = wb.create_sheet("Audit fattori emissivi")
        _build_emission_audit(ws_ef, ctx, lang=lang)

    # Imposta Piano come sheet attiva di default
    wb.active = wb.sheetnames.index(ws_piano.title)  # after i18n rename

    # === SNAPSHOT: protezione sheet (read-only) ===
    # Tutte le celle sono "locked" di default in Excel; abilitando
    # protection.sheet=True diventano effettivamente non editabili.
    # L'utente puo' comunque selezionare/copiare/stampare.
    # Per riprendere editing: Revisione > Rimuovi protezione foglio
    # (no password).
    if snapshot:
        for sname in wb.sheetnames:
            ws = wb[sname]
            ws.protection.sheet = True
            ws.protection.formatCells     = False
            ws.protection.formatColumns   = False
            ws.protection.formatRows      = False
            ws.protection.insertColumns   = False
            ws.protection.insertRows      = False
            ws.protection.deleteColumns   = False
            ws.protection.deleteRows      = False
            ws.protection.sort            = False
            ws.protection.autoFilter      = False
            ws.protection.pivotTables     = False
            ws.protection.selectLockedCells   = False  # consente click
            ws.protection.selectUnlockedCells = False

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_metaniq_xlsx_snapshot(ctx: dict) -> BytesIO:
    """Comodita': snapshot XLSX (valori statici, no formule).

    Equivalente a build_metaniq_xlsx(ctx, snapshot=True).
    """
    return build_metaniq_xlsx(ctx, snapshot=True)


# ============================================================
# Sheet 1 — Database feedstock (LAYOUT ORIZZONTALE)
# ============================================================
# Layout:
#   A1: "Parametro" | B1..N1: nomi biomasse (header)
#   A2: "Resa Nm3/t" | B2..N2: yield
#   A3: "eec"        | B3..N3: eec
#   A4: "esca"       | B4..N4: esca
#   A5: "etd"        | B5..N5: etd
#   A6: "ep"         | B6..N6: ep (linkato a Piano!B9)
#   A7: "e_total"    | B7..N7: =Bi3-Bi4+Bi5+Bi6 (formula)
#
# Le formule del Piano usano range ORIZZONTALI omogenei al
# bio_range del Piano (anch'esso orizzontale, una riga per mese).
# Cosi' SUMPRODUCT(C13:F13, Database!$B$2:$E$2) ha entrambe le
# matrici 1xN -> dot product corretto, no #VALORE.
# ============================================================
def _build_database(ws, ctx, lang='it'):
    feeds = ctx["active_feeds"]
    fdb   = ctx["FEEDSTOCK_DB"]
    n     = len(feeds)

    # === Riga 1: header (Parametro + nomi biomasse) ===
    c = ws.cell(row=1, column=1, value=_t("Parametro", lang))
    _style_header(c)
    for j, name in enumerate(feeds):
        c = ws.cell(row=1, column=2 + j, value=name)
        _style_header(c)
    ws.row_dimensions[1].height = 36

    # === Righe parametri (label sx + valori a destra) ===
    rows_def = [
        # (row, label, getter, fmt, style_amber)
        (2, "Resa Nm3/t",  lambda d, name: ctx.get("actual_yields", {}).get(name, float(d["yield"])), "0.0",     False),
        (3, "eec",         lambda d, name: ctx.get("actual_emissions", {}).get(name, {}).get("eec", float(d["eec"])),   "0.00",  False),
        (4, "esca",        lambda d, name: ctx.get("actual_emissions", {}).get(name, {}).get("esca", float(d["esca"])),  "0.00",  False),
        (5, "etd",         lambda d, name: ctx.get("actual_emissions", {}).get(name, {}).get("etd", float(d["etd"])),   "0.00",  False),
    ]
    for r, label, getter, fmt, _ in rows_def:
        # Label
        c_lbl = ws.cell(row=r, column=1, value=label)
        c_lbl.font = Font(bold=True, color=NAVY)
        c_lbl.fill = PatternFill("solid", fgColor=SLATE_50)
        c_lbl.alignment = Alignment(horizontal="left", indent=1)
        c_lbl.border = _border_thin()
        # Valori
        for j, name in enumerate(feeds):
            d = fdb[name]
            c_val = ws.cell(row=r, column=2 + j, value=getter(d, name))
            c_val.number_format = fmt
            c_val.fill = PatternFill("solid", fgColor=SLATE_50)
            c_val.alignment = Alignment(horizontal="right")
            c_val.border = _border_thin()

    # === Riga 6: ep linkato al Piano!B9 (master cell) ===
    c_lbl = ws.cell(row=6, column=1, value="ep (linkato Piano!B9)")
    c_lbl.font = Font(bold=True, color=NAVY)
    c_lbl.fill = PatternFill("solid", fgColor=SLATE_50)
    c_lbl.alignment = Alignment(horizontal="left", indent=1)
    c_lbl.border = _border_thin()
    for j in range(n):
        c_val = ws.cell(row=6, column=2 + j, value=f"='{_t('Piano mensile', lang)}'!$B$9")
        c_val.number_format = "0.00"
        c_val.fill = PatternFill("solid", fgColor=SLATE_50)
        c_val.alignment = Alignment(horizontal="right")
        c_val.border = _border_thin()

    # === Riga 7: e_total = eec - esca + etd + ep ===
    c_lbl = ws.cell(row=7, column=1, value="e_total")
    c_lbl.font = Font(bold=True, color=AMBER_DK)
    c_lbl.fill = PatternFill("solid", fgColor=AMBER_BG)
    c_lbl.alignment = Alignment(horizontal="left", indent=1)
    c_lbl.border = _border_thin()
    for j in range(n):
        col = 2 + j
        cl = get_column_letter(col)
        c_val = ws.cell(row=7, column=col,
                        value=f"={cl}3-{cl}4+{cl}5+{cl}6")
        c_val.number_format = "0.00"
        c_val.fill = PatternFill("solid", fgColor=AMBER_BG)
        c_val.alignment = Alignment(horizontal="right")
        c_val.font = Font(bold=True, color=AMBER_DK)
        c_val.border = _border_thin()

    # === Riga 8: Fonte del fattore emissivo eec (audit-trail UNI/TS) ===
    def _eec_origin_tag(_src):
        _s = _src or ""
        _sl = _s.lower()
        if _s.startswith("UNI-TS"):
            return "UNI/TS A.5"
        if "manure credit red iii" in _sl:
            return "RED III"
        if "jec" in _sl or "ktbl" in _sl:
            return "JEC/KTBL"
        if "gse" in _sl:
            return "GSE"
        if "ipcc" in _sl:
            return "IPCC"
        return "—"

    c_lbl = ws.cell(row=8, column=1, value=_t("Fonte fattore eec", lang))
    c_lbl.font = Font(bold=True, color=NAVY)
    c_lbl.fill = PatternFill("solid", fgColor=SLATE_50)
    c_lbl.alignment = Alignment(horizontal="left", indent=1)
    c_lbl.border = _border_thin()
    for j, name in enumerate(feeds):
        d = fdb[name]
        c_val = ws.cell(row=8, column=2 + j,
                        value=_eec_origin_tag(d.get("src", "")))
        c_val.fill = PatternFill("solid", fgColor=SLATE_50)
        c_val.alignment = Alignment(horizontal="center")
        c_val.font = Font(size=9, color=SLATE_500)
        c_val.border = _border_thin()

    # === Riga 9: Tier difendibilita' eec (A/B/C/D) ===
    def _eec_tier_code(_src, _eec):
        _s = (_src or "").strip()
        _sl = _s.lower()
        try:
            _e = float(_eec or 0.0)
        except (TypeError, ValueError):
            _e = 0.0
        if _s.startswith("UNI-TS"):
            return "A"
        if "manure credit red iii" in _sl or _e < 0:
            return "D"
        if _e == 0.0:
            return "B"
        return "C"

    c_lbl = ws.cell(row=9, column=1, value=_t("Tier difendibilità", lang))
    c_lbl.font = Font(bold=True, color=NAVY)
    c_lbl.fill = PatternFill("solid", fgColor=SLATE_50)
    c_lbl.alignment = Alignment(horizontal="left", indent=1)
    c_lbl.border = _border_thin()
    for j, name in enumerate(feeds):
        d = fdb[name]
        c_val = ws.cell(row=9, column=2 + j,
                        value=_eec_tier_code(d.get("src", ""), d.get("eec", 0.0)))
        c_val.fill = PatternFill("solid", fgColor=SLATE_50)
        c_val.alignment = Alignment(horizontal="center")
        c_val.font = Font(bold=True, size=9, color=NAVY)
        c_val.border = _border_thin()

    # === Caption finale ===
    last_r = 11
    end_col = 1 + n
    ws.merge_cells(start_row=last_r, start_column=1,
                   end_row=last_r, end_column=end_col)
    c = ws.cell(row=last_r, column=1,
                value=("Read-only · Layout orizzontale: ogni biomassa = una "
                       "colonna. e_total = eec - esca + etd + ep. "
                       "Modifica «ep» in «Piano mensile» cella B9 per "
                       "ricalcolare automaticamente la sostenibilita'. "
                       "Tier difendibilità eec: A=default normativo · "
                       "B=zero da regola (residuo/rifiuto) · "
                       "C=stima conservativa (letteratura, a sfavore) · "
                       "D=credito da dichiarazione fornitore."))
    c.font = Font(italic=True, size=9, color=SLATE_500)
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[last_r].height = 30

    # === Larghezze colonne ===
    ws.column_dimensions["A"].width = 26
    for j in range(n):
        ws.column_dimensions[get_column_letter(2 + j)].width = 16

    # Freeze pane: blocca header + colonna A
    ws.freeze_panes = ws["B2"]


# ============================================================
# Sheet 2 — Piano mensile (main, mode-aware + snapshot-aware)
# ============================================================
def _build_piano(ws, ctx, db_sheet_name, snapshot: bool = False, lang='it'):
    feeds   = ctx["active_feeds"]
    n_feed  = len(feeds)
    is_chp  = False  # legacy: app è mono-mode DM 2022 biometano
    # In modalita' SNAPSHOT i valori sono statici (no formule). Le celle
    # input (Ore, Biomasse) NON sono editabili (no fill amber).
    cell_fill_input = SLATE_50 if snapshot else AMBER_BG

    # Layout columns:
    #  A: Mese (read-only)
    #  B: Ore (editable)
    #  C..C+n-1: biomasse (editable)
    #
    # BIOMETANO (DM 2022, DM 2018):
    #  Sm3 lordi | Sm3 netti | MWh netti | e_w | Saving % | Sm3/h netti | Validita
    #
    # BIOGAS CHP (DM 2012, FER 2):
    #  Sm3 lordi (CH4 eq) | Sm3 netti (CH4 motore) | MWh CH4 netti |
    #  MWh el lordi | MWh el netti rete | MWh termici |
    #  e_w | Saving % | kW lordi (medio) | Validita
    bio_col_start = 3
    bio_col_end   = 2 + n_feed
    sm3_lordi_col = bio_col_end + 1
    sm3_netti_col = sm3_lordi_col + 1
    mwh_netti_col = sm3_netti_col + 1

    if is_chp:
        # Colonne aggiuntive per CHP
        mwh_el_lordo_col = mwh_netti_col + 1
        mwh_el_netto_col = mwh_el_lordo_col + 1
        mwh_th_col       = mwh_el_netto_col + 1
        e_w_col          = mwh_th_col + 1
        saving_col       = e_w_col + 1
        kw_lordi_col     = saving_col + 1   # kW elettrici LORDI medi
        valid_col        = kw_lordi_col + 1
    else:
        # Biometano: layout originale
        e_w_col          = mwh_netti_col + 1
        saving_col       = e_w_col + 1
        smch_col         = saving_col + 1   # Sm3/h netti immissione
        valid_col        = smch_col + 1
        # Placeholder per evitare NameError nei branch successivi
        mwh_el_lordo_col = mwh_el_netto_col = mwh_th_col = kw_lordi_col = None

    L = get_column_letter

    # === Title (row 1) ===
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=valid_col)
    c = ws.cell(
        row=1, column=1,
        value=("Metan.iQ — Monthly plan (snapshot)" if (snapshot and lang=="en") else "Metan.iQ — Monthly plan (editable)" if lang=="en" else "Metan.iQ — Piano mensile (snapshot)" if snapshot else "Metan.iQ — Piano mensile editabile"),
    )
    c.font = Font(bold=True, size=16, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    # === Subtitle (row 2) ===
    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2, end_column=valid_col)
    end_use = ctx.get("end_use", "")
    app_mode_label = ctx.get("APP_MODE_LABEL", "DM 2022")
    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    c = ws.cell(row=2, column=1,
                value=(
                    f"Modalita': {app_mode_label}  ·  "
                    f"Destinazione: {end_use}  ·  "
                    f"Generato il {now}"
                ))
    c.font = Font(italic=True, size=9, color=SLATE_500)
    c.alignment = Alignment(horizontal="left", indent=1)
    ws.row_dimensions[2].height = 18

    # === Helper banner (row 3) ===
    ws.merge_cells(start_row=3, start_column=1,
                   end_row=3, end_column=valid_col)
    if snapshot:
        banner_text = (
            "🔒 SNAPSHOT BLOCCATO — valori fotografati al download, "
            "celle in sola lettura. Per modificare e ricalcolare scarica "
            "il file «Excel modificabile» dall'app Metan.iQ."
        )
        banner_fg = SLATE_700
        banner_bg = SLATE_100
    else:
        banner_text = (
            "✏️ Modifica le celle GIALLE (Ore + Biomasse). "
            "Tutti i calcoli si aggiornano automaticamente."
        )
        banner_fg = AMBER_DK
        banner_bg = AMBER_BG
    c = ws.cell(row=3, column=1, value=banner_text)
    c.font = Font(bold=True, size=10, color=banner_fg)
    c.fill = PatternFill("solid", fgColor=banner_bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = _border_thin()
    ws.row_dimensions[3].height = 22

    # === Parametri impianto (rows 4-10 biometano, 4-13 CHP) ===
    aux_factor       = float(ctx.get("aux_factor", 1.29))
    comparator       = float(ctx.get("fossil_comparator", 80.0))
    ghg_threshold    = float(ctx.get("ghg_threshold", 0.65)) * 100
    plant_max_smch   = float(ctx.get("plant_net_smch", 300.0))
    ep_total         = float(ctx.get("ep_total", 0.0))
    nm3_to_mwh       = float(ctx.get("NM3_TO_MWH", 0.00979))
    # CHP-specific
    plant_kwe        = float(ctx.get("plant_kwe", 999.0))
    eta_el           = float(ctx.get("eta_el", 0.40))
    eta_th           = float(ctx.get("eta_th", 0.42))
    aux_el_pct       = float(ctx.get("aux_el_pct", 0.08)) * 100

    # Layout righe parametri (cell references in formulas):
    #   B5: aux_factor
    #   B6: comparator (CHP=183, biometano=80/94)
    #   B7: ghg_threshold (%)
    #   B8: plant_max  (CHP=plant_kwe in kWe LORDI; biometano=Sm3/h netti)
    #   B9: ep_total
    #   B10: PCI biometano (MWh/Sm3)
    #   --- CHP ONLY ---
    #   B11: eta_el
    #   B12: eta_th
    #   B13: aux_el_pct (%)
    common_max_label = (
        f"Potenza elettrica LORDA max (kWe targa)" if is_chp
        else f"Produzione netta max (Sm3/h)"
    )
    common_max_value = plant_kwe if is_chp else plant_max_smch
    common_max_fmt   = "0" if is_chp else "0.0"

    params = [
        ("PARAMETRI IMPIANTO (editabili)", None, None, True),
        ("aux_factor (netto -> lordo)",   aux_factor,    "0.000", False),
        ("Comparator fossile (gCO2/MJ)",  comparator,    "0",     False),
        ("Soglia saving GHG (%)",         ghg_threshold, "0.0",   False),
        (common_max_label,                common_max_value, common_max_fmt, False),
        ("ep totale (gCO2/MJ)",           ep_total,      "0.00",  False),
        ("PCI biometano (MWh/Sm3)",       nm3_to_mwh,    "0.00000", False),
    ]
    if is_chp:
        params.extend([
            ("Rendimento elettrico η_el", eta_el,    "0.000", False),
            ("Rendimento termico η_th",   eta_th,    "0.000", False),
            ("Autoconsumo ausiliari (% del lordo)", aux_el_pct, "0.0", False),
        ])
    for i, (label, value, fmt, is_header) in enumerate(params):
        r = 4 + i
        if is_header:
            ws.merge_cells(start_row=r, start_column=1,
                           end_row=r, end_column=2)
            c_lbl = ws.cell(row=r, column=1, value=label)
            c_lbl.font = Font(bold=True, color=WHITE, size=10)
            c_lbl.fill = PatternFill("solid", fgColor=NAVY_2)
            c_lbl.alignment = Alignment(horizontal="left", indent=1)
            c_lbl.border = _border_thin()
        else:
            c_lbl = ws.cell(row=r, column=1, value=label)
            c_lbl.font = Font(bold=True, color=SLATE_700)
            c_lbl.fill = PatternFill("solid", fgColor=SLATE_50)
            c_lbl.alignment = Alignment(horizontal="left", indent=1)
            c_lbl.border = _border_thin()

            c_val = ws.cell(row=r, column=2, value=value)
            c_val.number_format = fmt
            # Snapshot: read-only slate. Editable: amber (yellow).
            c_val.fill = PatternFill("solid", fgColor=cell_fill_input)
            c_val.font = Font(bold=True, color=NAVY)
            c_val.alignment = Alignment(horizontal="right")
            c_val.border = _border_thin()

    # === Empty row separatore ===
    # Per CHP la tabella inizia a row 15 (3 righe extra di params).
    # Per biometano resta a row 12 come prima.

    # === Header tabella (mode-aware row + columns) ===
    header_row = 15 if is_chp else 12
    if is_chp:
        # Etichette CHP (colonne aggiuntive: MWh el lordi/netti/termici, kW lordi)
        col_labels = [_t("Mese", lang), _t("Ore", lang)] + feeds + [
            "Sm3 CH4 lordi"    if lang=="it" else "Gross CH4 Sm³",
            "Sm3 CH4 al motore" if lang=="it" else "CH4 to engine",
            "MWh CH4 netti"    if lang=="it" else "Net CH4 MWh",
            "MWh el LORDI"     if lang=="it" else "Gross MWh el.",
            "MWh el netti rete" if lang=="it" else "Net MWh el. (grid)",
            "MWh termici"      if lang=="it" else "Thermal MWh",
            "e_w",
            "Saving %",
            "kW lordi (medio)" if lang=="it" else "Avg. gross kW",
            _t("Validità", lang),
        ]
    else:
        col_labels = [_t("Mese", lang), _t("Ore", lang)] + feeds + [
            _t("Sm³ lordi", lang), _t("Sm³ netti", lang), _t("MWh netti", lang),
            "e_w", "Saving %", _t("Sm³/h netti", lang), _t("Validità", lang),
        ]
    for col, h in enumerate(col_labels, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        _style_header(c)
    ws.row_dimensions[header_row].height = 42

    # === Riferimenti per formule (cell letters) ===
    aux_cell        = "$B$5"
    comparator_cell = "$B$6"
    threshold_cell  = "$B$7"
    max_prod_cell   = "$B$8"     # Sm3/h max biometano | kWe LORDI max CHP
    pci_cell        = "$B$10"
    # CHP-only cells
    eta_el_cell     = "$B$11"
    eta_th_cell     = "$B$12"
    aux_el_pct_cell = "$B$13"

    # Database layout ORIZZONTALE: yield in row 2, e_total in row 7
    # Per evitare problemi cross-version di SUMPRODUCT con range, usiamo
    # la forma ESPLICITA "somma di prodotti" - bulletproof in Excel,
    # LibreOffice, Numbers, Google Sheets, qualsiasi locale.
    # Esempio (4 biomasse):
    #   Sm3_lordi = C13*Database!$B$2 + D13*Database!$C$2 +
    #               E13*Database!$D$2 + F13*Database!$E$2

    def _sum_of_products(row_idx: int, db_row_idx: int) -> str:
        """Costruisce somma di prodotti esplicita per N_feed biomasse."""
        terms = []
        for j in range(n_feed):
            piano_col = L(bio_col_start + j)
            db_col    = L(2 + j)  # B per j=0, C per j=1, ...
            terms.append(
                f"{piano_col}{row_idx}*'{db_sheet_name}'!${db_col}${db_row_idx}"
            )
        return "+".join(terms)

    # === Dati 12 mesi ===
    months = ctx.get("MONTHS", [
        "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
        "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre",
    ])
    month_hours = ctx.get("MONTH_HOURS", [
        744, 672, 744, 720, 744, 720, 744, 744, 720, 744, 720, 744,
    ])
    initial_data = ctx.get("initial_data", {}) or {}

    # Per SNAPSHOT: leggiamo i valori computati gia' pronti dal df_res.
    # Mappiamo per Mese -> dict di tutte le colonne computate.
    snap_data = {}
    if snapshot:
        df = ctx.get("df_res")
        if df is not None:
            for _, row in df.iterrows():
                m_key = row.get("Mese")
                if m_key:
                    snap_data[m_key] = dict(row)

    first_data_row = header_row + 1
    last_data_row  = first_data_row + 11

    for i, (m, h) in enumerate(zip(months, month_hours)):
        r = first_data_row + i

        # A: Mese (read-only, navy header look-alike)
        c_mese = ws.cell(row=r, column=1, value=m)
        c_mese.font = Font(bold=True, color=WHITE, size=10)
        c_mese.fill = PatternFill("solid", fgColor=NAVY_2)
        c_mese.alignment = Alignment(horizontal="left", indent=1)
        c_mese.border = _border_thin()

        # B: Ore (editable in modificabile, read-only in snapshot)
        ore_default = h
        if snapshot and m in snap_data:
            ore_default = snap_data[m].get("Ore", h)
        elif m in initial_data:
            ore_default = initial_data[m].get("Ore", h)
        c_ore = ws.cell(row=r, column=2, value=int(ore_default))
        if snapshot:
            _style_readonly(c_ore)
        else:
            _style_editable(c_ore)
        c_ore.number_format = "0"

        # C..N: biomasse (editable in modificabile, read-only in snapshot)
        for j, name in enumerate(feeds):
            col = bio_col_start + j
            default = 0.0
            if snapshot and m in snap_data:
                default = float(snap_data[m].get(name, 0.0) or 0.0)
            elif m in initial_data:
                default = initial_data[m].get(name, 0.0)
            c_b = ws.cell(row=r, column=col, value=float(default))
            if snapshot:
                _style_readonly(c_b)
            else:
                _style_editable(c_b)
            c_b.number_format = "#,##0.0"

        # =====================================================
        # FORMULE LIVE (modificabile) o VALORI STATICI (snapshot)
        # =====================================================
        # Sm3 lordi
        if snapshot:
            v = float(snap_data.get(m, {}).get("Sm³ lordi") or 0)
            c = ws.cell(row=r, column=sm3_lordi_col, value=v)
        else:
            sop_yield = _sum_of_products(row_idx=r, db_row_idx=2)
            c = ws.cell(row=r, column=sm3_lordi_col,
                        value=f"=IFERROR({sop_yield},0)")
        c.number_format = "#,##0"
        _style_readonly(c)

        # Sm3 netti
        sm3_lordi_letter = L(sm3_lordi_col)
        if snapshot:
            v = float(snap_data.get(m, {}).get("Sm³ netti") or 0)
            c = ws.cell(row=r, column=sm3_netti_col, value=v)
        else:
            c = ws.cell(row=r, column=sm3_netti_col,
                        value=f"=IFERROR({sm3_lordi_letter}{r}/{aux_cell},0)")
        c.number_format = "#,##0"
        _style_readonly(c)

        # MWh netti
        sm3_netti_letter = L(sm3_netti_col)
        if snapshot:
            v = float(snap_data.get(m, {}).get("MWh netti") or 0)
            c = ws.cell(row=r, column=mwh_netti_col, value=v)
        else:
            c = ws.cell(row=r, column=mwh_netti_col,
                        value=f"={sm3_netti_letter}{r}*{pci_cell}")
        c.number_format = "#,##0.0"
        _style_readonly(c)

        # e_w (gCO2/MJ)
        if snapshot:
            v = float(snap_data.get(m, {}).get("GHG (gCO₂/MJ)") or 0)
            c = ws.cell(row=r, column=e_w_col, value=v)
        else:
            num_terms = []
            for j in range(n_feed):
                piano_col = L(bio_col_start + j)
                db_col    = L(2 + j)
                num_terms.append(
                    f"{piano_col}{r}*'{db_sheet_name}'!${db_col}$2*"
                    f"'{db_sheet_name}'!${db_col}$7"
                )
            sop_num = "+".join(num_terms)
            sop_yield = _sum_of_products(row_idx=r, db_row_idx=2)
            c = ws.cell(row=r, column=e_w_col,
                        value=f"=IFERROR(({sop_num})/({sop_yield}),0)")
        c.number_format = "0.00"
        _style_readonly(c)

        # === CHP-only colonne energetiche ===
        if is_chp:
            mwh_netti_letter = L(mwh_netti_col)
            # MWh el LORDI = MWh CH4 × η_el
            if snapshot:
                v = float(snap_data.get(m, {}).get("MWh elettrici lordi") or 0)
                c = ws.cell(row=r, column=mwh_el_lordo_col, value=v)
            else:
                c = ws.cell(row=r, column=mwh_el_lordo_col,
                            value=f"={mwh_netti_letter}{r}*{eta_el_cell}")
            c.number_format = "#,##0.0"
            _style_readonly(c)

            # MWh el NETTI rete = lordi × (1 - aux%/100)
            mwh_el_lordo_letter = L(mwh_el_lordo_col)
            if snapshot:
                v = float(snap_data.get(m, {}).get("MWh elettrici netti") or 0)
                c = ws.cell(row=r, column=mwh_el_netto_col, value=v)
            else:
                c = ws.cell(row=r, column=mwh_el_netto_col,
                            value=(f"={mwh_el_lordo_letter}{r}*"
                                   f"(1-{aux_el_pct_cell}/100)"))
            c.number_format = "#,##0.0"
            _style_readonly(c)

            # MWh termici = MWh CH4 × η_th
            if snapshot:
                v = float(snap_data.get(m, {}).get("MWh termici") or 0)
                c = ws.cell(row=r, column=mwh_th_col, value=v)
            else:
                c = ws.cell(row=r, column=mwh_th_col,
                            value=f"={mwh_netti_letter}{r}*{eta_th_cell}")
            c.number_format = "#,##0.0"
            _style_readonly(c)

        # === Saving % ===
        e_w_letter = L(e_w_col)
        if snapshot:
            v = float(snap_data.get(m, {}).get("Saving %") or 0)
            c = ws.cell(row=r, column=saving_col, value=v)
        else:
            c = ws.cell(row=r, column=saving_col,
                        value=(f"=IFERROR(({comparator_cell}-{e_w_letter}{r})"
                               f"/{comparator_cell}*100,0)"))
        c.number_format = "0.0\"%\""
        _style_readonly(c)

        # === Production check column (mode-aware) ===
        if is_chp:
            # kW lordi medi sull'ora = MWh el lordi × 1000 / Ore
            mwh_el_lordo_letter = L(mwh_el_lordo_col)
            if snapshot:
                # Calcoliamo da MWh el lordi e Ore (gia' nel snap)
                _ml = float(snap_data.get(m, {}).get("MWh elettrici lordi") or 0)
                _ore = float(snap_data.get(m, {}).get("Ore") or h)
                v = (_ml * 1000.0 / _ore) if _ore > 0 else 0.0
                c = ws.cell(row=r, column=kw_lordi_col, value=v)
            else:
                c = ws.cell(row=r, column=kw_lordi_col,
                            value=(f"=IFERROR({mwh_el_lordo_letter}{r}*1000"
                                   f"/B{r},0)"))
            c.number_format = "#,##0"
            _style_readonly(c)
            prod_check_letter = L(kw_lordi_col)
        else:
            # Biometano: Sm3/h netti = Sm3 netti / Ore
            if snapshot:
                v = float(snap_data.get(m, {}).get("Sm³/h netti") or 0)
                c = ws.cell(row=r, column=smch_col, value=v)
            else:
                c = ws.cell(row=r, column=smch_col,
                            value=f"=IFERROR({sm3_netti_letter}{r}/B{r},0)")
            c.number_format = "0.0"
            _style_readonly(c)
            prod_check_letter = L(smch_col)

        # === Validita ===
        saving_letter = L(saving_col)
        if snapshot:
            # df_res ha "Validita" con emoji ✅/❌. Normalizziamo a OK/KO.
            v_str = str(snap_data.get(m, {}).get("Validità") or "")
            if "✅" in v_str:
                v_clean = "OK"
            elif "❌" in v_str:
                v_clean = "KO"
            elif v_str.upper().startswith("OK"):
                v_clean = "OK"
            else:
                v_clean = "KO"
            c = ws.cell(row=r, column=valid_col, value=v_clean)
        else:
            c = ws.cell(row=r, column=valid_col,
                        value=(f'=IF(AND({saving_letter}{r}>={threshold_cell},'
                               f'{prod_check_letter}{r}<={max_prod_cell}),"OK","KO")'))
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border_thin()

    # === Riga TOTALE (row 25) ===
    tot_row = last_data_row + 1
    c_tot_label = ws.cell(row=tot_row, column=1, value="TOTALE/MEDIA")
    _style_total(c_tot_label)
    c_tot_label.alignment = Alignment(horizontal="left", indent=1)

    # Ore tot
    c = ws.cell(row=tot_row, column=2,
                value=f"=SUM(B{first_data_row}:B{last_data_row})")
    _style_total(c); c.number_format = "0"

    # Biomasse tot
    for j in range(n_feed):
        col = bio_col_start + j
        cl = L(col)
        c = ws.cell(row=tot_row, column=col,
                    value=f"=SUM({cl}{first_data_row}:{cl}{last_data_row})")
        _style_total(c); c.number_format = "#,##0"

    # Sm3 lordi/netti/MWh tot
    sum_cols = [sm3_lordi_col, sm3_netti_col, mwh_netti_col]
    if is_chp:
        sum_cols += [mwh_el_lordo_col, mwh_el_netto_col, mwh_th_col]
    for col in sum_cols:
        cl = L(col)
        c = ws.cell(row=tot_row, column=col,
                    value=f"=SUM({cl}{first_data_row}:{cl}{last_data_row})")
        _style_total(c)
        c.number_format = (
            "#,##0" if col in (sm3_lordi_col, sm3_netti_col)
            else "#,##0.0"
        )

    # e_w medio (weighted by Sm3 netti)
    cl_ew     = L(e_w_col)
    cl_smnett = L(sm3_netti_col)
    c = ws.cell(row=tot_row, column=e_w_col,
                value=(f"=IFERROR(SUMPRODUCT("
                       f"{cl_smnett}{first_data_row}:{cl_smnett}{last_data_row},"
                       f"{cl_ew}{first_data_row}:{cl_ew}{last_data_row})"
                       f"/SUM({cl_smnett}{first_data_row}:{cl_smnett}{last_data_row}),0)"))
    _style_total(c); c.number_format = "0.00"

    # Saving medio (weighted)
    cl_sav = L(saving_col)
    c = ws.cell(row=tot_row, column=saving_col,
                value=(f"=IFERROR(SUMPRODUCT("
                       f"{cl_smnett}{first_data_row}:{cl_smnett}{last_data_row},"
                       f"{cl_sav}{first_data_row}:{cl_sav}{last_data_row})"
                       f"/SUM({cl_smnett}{first_data_row}:{cl_smnett}{last_data_row}),0)"))
    _style_total(c); c.number_format = "0.0\"%\""

    # Production check medio (Sm3/h biometano | kW lordi CHP)
    if is_chp:
        cl_prod = L(kw_lordi_col)
        prod_fmt = "#,##0"
    else:
        cl_prod = L(smch_col)
        prod_fmt = "0.0"
    c = ws.cell(row=tot_row, column=(kw_lordi_col if is_chp else smch_col),
                value=(f"=IFERROR(AVERAGE("
                       f"{cl_prod}{first_data_row}:{cl_prod}{last_data_row}),0)"))
    _style_total(c); c.number_format = prod_fmt

    # Validita: count "OK" / 12
    cl_val = L(valid_col)
    c = ws.cell(row=tot_row, column=valid_col,
                value=(f'=COUNTIF({cl_val}{first_data_row}:{cl_val}{last_data_row},'
                       f'"OK")&"/12"'))
    _style_total(c); c.alignment = Alignment(horizontal="center")

    # === Conditional formatting ===
    # Saving %: rosso se < soglia, verde se >=
    sav_range = f"{cl_sav}{first_data_row}:{cl_sav}{last_data_row}"
    rule_red = CellIsRule(
        operator="lessThan", formula=[threshold_cell],
        fill=PatternFill("solid", fgColor=RED_BG),
        font=Font(color=RED_FG, bold=True),
    )
    rule_grn = CellIsRule(
        operator="greaterThanOrEqual", formula=[threshold_cell],
        fill=PatternFill("solid", fgColor=EMERALD_BG),
        font=Font(color=EMERALD_FG, bold=True),
    )
    ws.conditional_formatting.add(sav_range, rule_red)
    ws.conditional_formatting.add(sav_range, rule_grn)

    # Validita: OK = verde, KO = rosso
    val_range = f"{cl_val}{first_data_row}:{cl_val}{last_data_row}"
    rule_ok = CellIsRule(
        operator="equal", formula=['"OK"'],
        fill=PatternFill("solid", fgColor=EMERALD_BG),
        font=Font(color=EMERALD_FG, bold=True),
    )
    rule_ko = CellIsRule(
        operator="equal", formula=['"KO"'],
        fill=PatternFill("solid", fgColor=RED_BG),
        font=Font(color=RED_FG, bold=True),
    )
    ws.conditional_formatting.add(val_range, rule_ok)
    ws.conditional_formatting.add(val_range, rule_ko)

    # Production check rosso se > max (Sm3/h biometano | kW lordi CHP)
    prod_range = f"{cl_prod}{first_data_row}:{cl_prod}{last_data_row}"
    rule_prod_ko = CellIsRule(
        operator="greaterThan", formula=[max_prod_cell],
        fill=PatternFill("solid", fgColor=RED_BG),
        font=Font(color=RED_FG),
    )
    ws.conditional_formatting.add(prod_range, rule_prod_ko)

    # === Larghezze colonne ===
    ws.column_dimensions["A"].width = 14   # Mese
    ws.column_dimensions["B"].width = 8    # Ore
    for j in range(n_feed):
        ws.column_dimensions[L(bio_col_start + j)].width = 14
    ws.column_dimensions[L(sm3_lordi_col)].width = 13
    ws.column_dimensions[L(sm3_netti_col)].width = 13
    ws.column_dimensions[L(mwh_netti_col)].width = 12
    ws.column_dimensions[L(e_w_col)].width      = 9
    ws.column_dimensions[L(saving_col)].width   = 10
    ws.column_dimensions[L(valid_col)].width    = 10
    if is_chp:
        ws.column_dimensions[L(mwh_el_lordo_col)].width = 13
        ws.column_dimensions[L(mwh_el_netto_col)].width = 14
        ws.column_dimensions[L(mwh_th_col)].width       = 12
        ws.column_dimensions[L(kw_lordi_col)].width     = 13
    else:
        ws.column_dimensions[L(smch_col)].width = 12

    # === Freeze panes (header + Mese/Ore visibili) ===
    ws.freeze_panes = ws[f"C{first_data_row}"]

    # === Footer note ===
    note_row = tot_row + 2
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=valid_col)
    note = (
        "Generato da Metan.iQ - Decision Intelligence Platform per "
        "biometano e biogas CHP. Modello GHG conforme RED III All. V/VI. "
        "Le celle gialle sono editabili: ogni modifica ricalcola "
        "in tempo reale produzione, sostenibilita' e validita'."
    )
    c = ws.cell(row=note_row, column=1, value=note)
    c.font = Font(italic=True, size=8, color=SLATE_500)
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[note_row].height = 30


# ============================================================
# Sheet 3 — Sintesi annuale (mode-aware)
# ============================================================
def _build_summary(ws, ctx, piano_sheet_name, lang='it'):
    feeds  = ctx["active_feeds"]
    n_feed = len(feeds)
    is_chp = False  # legacy: app è mono-mode DM 2022 biometano

    L = get_column_letter
    bio_col_start = 3
    bio_col_end   = 2 + n_feed
    sm3_lordi_col = bio_col_end + 1
    sm3_netti_col = sm3_lordi_col + 1
    mwh_netti_col = sm3_netti_col + 1

    if is_chp:
        # Layout CHP: cols Sm3 lordi, Sm3 netti, MWh CH4, MWh el lordi,
        #             MWh el netti rete, MWh termici, e_w, Saving, kW lordi, Validita
        mwh_el_lordo_col = mwh_netti_col + 1
        mwh_el_netto_col = mwh_el_lordo_col + 1
        mwh_th_col       = mwh_el_netto_col + 1
        e_w_col          = mwh_th_col + 1
        saving_col       = e_w_col + 1
        kw_lordi_col     = saving_col + 1
        valid_col        = kw_lordi_col + 1
        first_data_row = 16  # CHP ha header a row 15
        last_data_row  = 27
    else:
        # Biometano: cols Sm3 lordi, Sm3 netti, MWh netti, e_w, Saving, Sm3/h, Validita
        e_w_col      = mwh_netti_col + 1
        saving_col   = e_w_col + 1
        smch_col     = saving_col + 1
        valid_col    = smch_col + 1
        first_data_row = 13
        last_data_row  = 24

    p = piano_sheet_name

    # === Title ===
    ws.merge_cells("A1:D1")
    c = ws.cell(row=1, column=1, value=_t("Metan.iQ — Sintesi annuale", lang) if lang=="en" else "Metan.iQ — Sintesi annuale")
    c.font = Font(bold=True, size=14, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    c = ws.cell(row=2, column=1,
                value=("Updated in real-time from 'Monthly plan' sheet. Edit hours/feedstocks there to see KPIs change." if lang=="en" else "Aggiornata in tempo reale dalla sheet «Piano mensile». Modifica ore/biomasse li' per vedere i KPI cambiare.")
                )
    c.font = Font(italic=True, size=9, color=SLATE_500)
    c.alignment = Alignment(horizontal="left", indent=1)

    # === KPI block (mode-aware) ===
    cl_sav  = L(saving_col)
    cl_val  = L(valid_col)
    cl_mwh  = L(mwh_netti_col)
    # Threshold cell location nel Piano (CHP=$B$7, biometano=$B$7) - same
    threshold_ref = f"'{p}'!$B$7"

    kpi_common = [
        ("Tot. biomasse (t/anno)",
         f"=SUMPRODUCT(('{p}'!{L(bio_col_start)}{first_data_row}:"
         f"{L(bio_col_end)}{last_data_row}))",
         "#,##0"),
    ]

    if is_chp:
        cl_lordo = L(mwh_el_lordo_col)
        cl_netto = L(mwh_el_netto_col)
        cl_th    = L(mwh_th_col)
        cl_kw    = L(kw_lordi_col)
        kpi = kpi_common + [
            ("MWh CH4 al motore (anno)",
             f"=SUM('{p}'!{cl_mwh}{first_data_row}:{cl_mwh}{last_data_row})",
             "#,##0.0"),
            ("MWh el LORDI (anno)",
             f"=SUM('{p}'!{cl_lordo}{first_data_row}:{cl_lordo}{last_data_row})",
             "#,##0.0"),
            ("⚡ MWh el NETTI rete (anno)",
             f"=SUM('{p}'!{cl_netto}{first_data_row}:{cl_netto}{last_data_row})",
             "#,##0.0"),
            ("🔥 MWh termici (anno)",
             f"=SUM('{p}'!{cl_th}{first_data_row}:{cl_th}{last_data_row})",
             "#,##0.0"),
            ("kW lordi medi (anno)",
             f"=IFERROR(AVERAGE('{p}'!{cl_kw}{first_data_row}:"
             f"{cl_kw}{last_data_row}),0)",
             "#,##0"),
            ("Saving GHG medio (%)",
             f"=IFERROR(SUMPRODUCT('{p}'!{cl_mwh}{first_data_row}:"
             f"{cl_mwh}{last_data_row},"
             f"'{p}'!{cl_sav}{first_data_row}:{cl_sav}{last_data_row})"
             f"/SUM('{p}'!{cl_mwh}{first_data_row}:"
             f"{cl_mwh}{last_data_row}),0)",
             "0.0\"%\""),
            ("Soglia RED III (%)",       f"={threshold_ref}", "0.0\"%\""),
            ("Mesi validi (saving + kW lordi)",
             f'=COUNTIF(\'{p}\'!{cl_val}{first_data_row}:'
             f'{cl_val}{last_data_row},"OK")&"/12"',
             None),
        ]
    else:
        cl_smnett = L(sm3_netti_col)
        kpi = kpi_common + [
            ("Sm3 netti totali (anno)",
             f"=SUM('{p}'!{cl_smnett}{first_data_row}:"
             f"{cl_smnett}{last_data_row})",
             "#,##0"),
            ("MWh netti totali (anno)",
             f"=SUM('{p}'!{cl_mwh}{first_data_row}:{cl_mwh}{last_data_row})",
             "#,##0.0"),
            ("Saving GHG medio (%)",
             f"=IFERROR(SUMPRODUCT('{p}'!{cl_smnett}{first_data_row}:"
             f"{cl_smnett}{last_data_row},"
             f"'{p}'!{cl_sav}{first_data_row}:{cl_sav}{last_data_row})"
             f"/SUM('{p}'!{cl_smnett}{first_data_row}:"
             f"{cl_smnett}{last_data_row}),0)",
             "0.0\"%\""),
            ("Soglia RED III (%)",       f"={threshold_ref}", "0.0\"%\""),
            ("Mesi validi (saving + produzione)",
             f'=COUNTIF(\'{p}\'!{cl_val}{first_data_row}:'
             f'{cl_val}{last_data_row},"OK")&"/12"',
             None),
        ]
    for i, (lbl, formula, fmt) in enumerate(kpi):
        r = 4 + i
        c_lbl = ws.cell(row=r, column=1, value=lbl)
        c_lbl.font = Font(bold=True, color=SLATE_700)
        c_lbl.fill = PatternFill("solid", fgColor=SLATE_50)
        c_lbl.alignment = Alignment(horizontal="left", indent=1)
        c_lbl.border = _border_thin()
        ws.merge_cells(start_row=r, start_column=1,
                       end_row=r, end_column=2)

        c_val = ws.cell(row=r, column=3, value=formula)
        c_val.font = Font(bold=True, color=NAVY, size=12)
        c_val.fill = PatternFill("solid", fgColor=AMBER_BG)
        c_val.alignment = Alignment(horizontal="right", vertical="center")
        c_val.border = _border_thin()
        if fmt:
            c_val.number_format = fmt
        ws.merge_cells(start_row=r, start_column=3,
                       end_row=r, end_column=4)
        ws.row_dimensions[r].height = 24

    # === Mix biomasse (annuale) ===
    mix_row_hdr = 4 + len(kpi) + 2
    ws.merge_cells(start_row=mix_row_hdr, start_column=1,
                   end_row=mix_row_hdr, end_column=4)
    c = ws.cell(row=mix_row_hdr, column=1, value="MIX BIOMASSE ANNUALE")
    c.font = Font(bold=True, color=WHITE, size=11)
    c.fill = PatternFill("solid", fgColor=NAVY_2)
    c.alignment = Alignment(horizontal="left", indent=1)
    c.border = _border_thin()
    ws.row_dimensions[mix_row_hdr].height = 22

    headers = ["Biomassa", "t/anno", "Quota %", "MWh CH4 equiv."]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=mix_row_hdr + 1, column=col, value=h)
        _style_header(c)

    for j, name in enumerate(feeds):
        r = mix_row_hdr + 2 + j
        col = bio_col_start + j
        cl = L(col)
        # Nome
        ws.cell(row=r, column=1, value=name).font = Font(bold=True)
        # t/anno
        c = ws.cell(row=r, column=2,
                    value=f"=SUM('{p}'!{cl}{first_data_row}:"
                          f"{cl}{last_data_row})")
        c.number_format = "#,##0"; _style_readonly(c)
        # Quota %
        tot_t_formula = (
            f"SUMPRODUCT('{p}'!{L(bio_col_start)}{first_data_row}:"
            f"{L(bio_col_end)}{last_data_row})"
        )
        c = ws.cell(row=r, column=3,
                    value=f"=IFERROR(B{r}/{tot_t_formula}*100,0)")
        c.number_format = "0.0\"%\""; _style_readonly(c)
        # MWh CH4 equiv = t × yield × PCI / aux_factor
        # NB Database orizzontale: yield in row 2, biomassa j -> col (2+j)
        db_yield_col_letter = get_column_letter(2 + j)
        c = ws.cell(row=r, column=4,
                    value=f"=B{r}*'{_t('Database feedstock', lang)}'!${db_yield_col_letter}$2*"
                          f"'{p}'!$B$10/'{p}'!$B$5")
        c.number_format = "#,##0.0"; _style_readonly(c)

    # Larghezze
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18


# ============================================================
# Sheet 4 — Business Plan (Pro Forma 15 anni, mode-aware)
# ============================================================


# ============================================================
# Sheet 5 — Audit Rese BMT (override certificati vs tabella standard)
# ============================================================
def _build_bmt_audit(ws, ctx, lang="it"):
    """Costruisce il foglio audit rese BMT.

    Mostra, per ogni biomassa attiva:
      - resa standard (tabella interna UNI-TS / JEC v5)
      - resa usata nei calcoli
      - unita' (Sm3 biometano/t)
      - origine resa (tabella standard | BMT certificato laboratorio)
      - certificato (filename)
      - laboratorio
      - data certificato
      - riferimento campione

    Se non ci sono override BMT, il foglio mostra comunque tutte le
    biomasse attive con origine "tabella standard" — utile come traccia
    di compliance per il consulente / auditor.
    """
    audit_rows = ctx.get("yield_audit_rows", []) or []

    # Title
    ws.merge_cells("A1:I1")
    c = ws.cell(row=1, column=1,
                value=("Metan.iQ - BMT Yield Audit (certified vs standard)"
                       if lang == "en" else
                       "Metan.iQ - Audit Rese BMT (certificate vs standard)"))
    c.font = Font(bold=True, size=14, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:I2")
    c = ws.cell(row=2, column=1, value=(
        "For each active feedstock: standard yield, BMT-certified yield "
        "(if active), origin, lab metadata. Standard table is never modified."
        if lang == "en" else
        "Per ogni biomassa attiva: resa standard, resa BMT certificata "
        "(se attiva), origine, metadati laboratorio. La tabella standard "
        "non viene mai modificata."
    ))
    c.font = Font(italic=True, size=9, color=SLATE_500)
    c.alignment = Alignment(horizontal="left", indent=1)

    # Header
    if lang == "en":
        headers = [
            "Feedstock", "Standard yield", "Used yield", "Unit",
            "Yield source", "Certificate", "Laboratory",
            "Certificate date", "Sample ref",
        ]
    else:
        headers = [
            "Biomassa", "Resa standard", "Resa usata", "Unita'",
            "Origine resa", "Certificato", "Laboratorio",
            "Data certificato", "Riferimento campione",
        ]
    header_row = 4
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        _style_header(c)
    ws.row_dimensions[header_row].height = 30

    # Data rows
    if not audit_rows:
        ws.merge_cells(start_row=header_row + 1, start_column=1,
                       end_row=header_row + 1, end_column=9)
        c = ws.cell(row=header_row + 1, column=1,
                    value=("No active feedstock - select biomasses in the app."
                           if lang == "en" else
                           "Nessuna biomassa attiva - seleziona biomasse nell'app."))
        c.font = Font(italic=True, color=SLATE_500)
        c.alignment = Alignment(horizontal="center")
    else:
        for i, r in enumerate(audit_rows, start=header_row + 1):
            cells = [
                r.get("Biomassa", ""),
                float(r.get("Resa standard", 0.0)),
                float(r.get("Resa usata", 0.0)),
                r.get("Unita'", r.get("Unita", r.get("Unitá", r.get("Unit", "")))) or r.get("Unità", "Sm3 biometano/t"),
                r.get("Origine resa", ""),
                r.get("Certificato", "-"),
                r.get("Laboratorio", "-"),
                r.get("Data certificato", "-"),
                r.get("Riferimento campione", "-"),
            ]
            # Recover unit field robustly: the build_yield_audit_row uses key "Unita'" with combining mark
            unit_val = ""
            for k, v in r.items():
                if "nit" in k.lower():
                    unit_val = v
                    break
            cells[3] = unit_val or "Sm3 biometano/t"

            for col, val in enumerate(cells, start=1):
                c = ws.cell(row=i, column=col, value=val)
                _style_readonly(c)
                c.alignment = Alignment(
                    horizontal="left" if col in (1, 4, 5, 6, 7, 8, 9)
                    else "right",
                    vertical="center",
                )
            # Numeric format for yield columns
            ws.cell(row=i, column=2).number_format = "0.0"
            ws.cell(row=i, column=3).number_format = "0.0"

            # Highlight rows with BMT override (amber background on origin column)
            origin = str(r.get("Origine resa", ""))
            if "BMT" in origin or "certif" in origin.lower():
                for col in (3, 5):  # Resa usata + Origine resa
                    cell = ws.cell(row=i, column=col)
                    cell.fill = PatternFill("solid", fgColor=AMBER_BG)
                    cell.font = Font(bold=True, color=AMBER_DK)

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 36
    ws.column_dimensions["F"].width = 24
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 22

    ws.freeze_panes = ws[f"A{header_row + 1}"]

    # Footer note
    note_row = header_row + len(audit_rows) + 2
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=9)
    note = (
        "BMT (Biochemical Methane Test) override: certified lab yield "
        "replaces the internal standard table for the specific feedstock "
        "only. Standard table is never modified. Warning issued if BMT "
        "deviates more than +/-30% from standard."
        if lang == "en" else
        "BMT (Biochemical Methane Test) override: la resa certificata "
        "in laboratorio sostituisce la tabella standard interna SOLO "
        "per la specifica biomassa. La tabella standard non viene mai "
        "modificata. Warning emesso se il BMT differisce oltre +/-30% "
        "dal valore standard."
    )
    c = ws.cell(row=note_row, column=1, value=note)
    c.font = Font(italic=True, size=8, color=SLATE_500)
    c.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[note_row].height = 30


# ============================================================
# Sheet 6 - Audit Fattori Emissivi (override REALI vs standard)
# ============================================================
# ============================================================
# Sheet 5 - Audit Fattori Emissivi (override REALI vs standard)
# ============================================================
def _build_emission_audit(ws, ctx, lang="it"):
    """Foglio audit fattori emissivi reali da relazione tecnica.

    Mostra, per ogni biomassa attiva: fattori standard (eec/esca/etd/ep),
    fattori usati nei calcoli, scostamento % (solo se override attivo),
    crediti extra, e_total, origine, metadati relazione.

    Le righe con override REALE attivo sono evidenziate in giallo amber
    sulla colonna Origine + colonne fattori usati. La tabella standard
    NON viene mai modificata: questo foglio serve solo come traccia
    di compliance.
    """
    audit_rows = ctx.get("emission_audit_rows", []) or []

    # Title
    # Headers: 20 columns (A..T) — merge title across full width
    ws.merge_cells("A1:T1")
    c = ws.cell(row=1, column=1,
                value=("Metan.iQ - Real Emission Factors Audit"
                       if lang == "en" else
                       "Metan.iQ - Audit Fattori Emissivi (relazione tecnica vs standard)"))
    c.font = Font(bold=True, size=14, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:T2")
    c = ws.cell(row=2, column=1, value=(
        "Formula: e_total = eec + etd + ep - esca - extra_credits. "
        "The standard table is never permanently modified by overrides."
        if lang == "en" else
        "Formula: e_total = eec + etd + ep - esca - crediti_extra. "
        "La tabella standard NON viene mai modificata permanentemente."
    ))
    c.font = Font(italic=True, size=9, color=SLATE_500)
    c.alignment = Alignment(horizontal="left", indent=1)

    # Header
    if lang == "en":
        headers = [
            "Feedstock", "Source", "eec std", "eec used", "eec dev.%",
            "esca std", "esca used", "etd std", "etd used",
            "ep std", "ep used", "extra credits", "e_total",
            "Tech. report", "Author", "Company", "Date",
            "Plant ref.", "Sample ref.", "Methodology notes",
        ]
    else:
        headers = [
            "Biomassa", "Origine", "eec std", "eec usato", "eec scost.%",
            "esca std", "esca usato", "etd std", "etd usato",
            "ep std", "ep usato", "Crediti extra", "e_total",
            "Relazione", "Autore", "Societa'", "Data",
            "Impianto", "Rif. campione", "Note metodol.",
        ]
    header_row = 4
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        _style_header(c)
    ws.row_dimensions[header_row].height = 36

    # Data rows
    if not audit_rows:
        ws.merge_cells(start_row=header_row + 1, start_column=1,
                       end_row=header_row + 1, end_column=len(headers))
        c = ws.cell(row=header_row + 1, column=1,
                    value=("No active feedstock." if lang == "en"
                           else "Nessuna biomassa attiva."))
        c.font = Font(italic=True, color=SLATE_500)
        c.alignment = Alignment(horizontal="center")
    else:
        for i, r in enumerate(audit_rows, start=header_row + 1):
            cells = [
                r.get("Biomassa", ""),
                r.get("Origine fattori", ""),
                float(r.get("eec standard", 0.0)),
                float(r.get("eec usato", 0.0)),
                r.get("eec scost. %", ""),
                float(r.get("esca standard", 0.0)),
                float(r.get("esca usato", 0.0)),
                float(r.get("etd standard", 0.0)),
                float(r.get("etd usato", 0.0)),
                float(r.get("ep standard", 0.0)),
                float(r.get("ep usato", 0.0)),
                float(r.get("Crediti extra", 0.0)),
                float(r.get("e_total", 0.0)),
                r.get("Relazione tecnica", "-"),
                r.get("Autore", "-"),
                r.get("Societa'", "-"),
                r.get("Data relazione", "-"),
                r.get("Impianto rif.", "-"),
                r.get("Riferimento campione", "-"),
                r.get("Note metodologiche", "-"),
            ]
            for col, val in enumerate(cells, start=1):
                c = ws.cell(row=i, column=col, value=val)
                _style_readonly(c)
                # Numeric formatting per columns
                if col in (3, 4):
                    c.number_format = "+0.00;-0.00"
                elif col in (6, 7, 8, 9, 10, 11, 12):
                    c.number_format = "0.00"
                elif col == 13:  # e_total
                    c.number_format = "+0.00;-0.00"
                c.alignment = Alignment(
                    horizontal="left" if col in (1, 2, 5, 14, 15, 16, 17, 18, 19, 20)
                    else "right",
                    vertical="center",
                )

            # Highlight rows with REAL override
            origin = str(r.get("Origine fattori", ""))
            if "Relazione" in origin or "elaz" in origin.lower():
                for col in (2, 4, 7, 9, 11, 12, 13):
                    cell = ws.cell(row=i, column=col)
                    cell.fill = PatternFill("solid", fgColor=AMBER_BG)
                    cell.font = Font(bold=True, color=AMBER_DK)

    # Column widths
    widths = [
        ("A", 26), ("B", 28), ("C", 11), ("D", 11), ("E", 12),
        ("F", 11), ("G", 11), ("H", 11), ("I", 11),
        ("J", 11), ("K", 11), ("L", 13), ("M", 11),
        ("N", 22), ("O", 18), ("P", 18), ("Q", 12),
        ("R", 22), ("S", 18), ("T", 28),
    ]
    for col, w in widths:
        ws.column_dimensions[col].width = w
    ws.freeze_panes = ws[f"A{header_row + 1}"]

    # Footer
    note_row = header_row + max(len(audit_rows), 1) + 2
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=len(headers))
    note = (
        "Real Emission Factors override: technical-report values replace "
        "the internal standard table for the specific feedstock only. "
        "Standard table is never modified. Warning issued if a real factor "
        "deviates more than +/-30% from standard. Validation rules: report "
        "file mandatory (PDF/DOCX/XLSX/CSV/JPG/PNG); numeric finite values; "
        "esca/etd/ep/extra_credits >= 0; complete metadata; biomass-bound."
        if lang == "en" else
        "Override Fattori Emissivi Reali: i valori dichiarati nella relazione "
        "tecnica sostituiscono la tabella standard interna SOLO per la "
        "biomassa specifica. La tabella standard non viene mai modificata. "
        "Warning se un fattore reale scosta oltre +/-30% dallo standard. "
        "Regole di validazione: relazione tecnica obbligatoria "
        "(PDF/DOCX/XLSX/CSV/JPG/PNG); valori numerici finiti; "
        "esca/etd/ep/crediti_extra >= 0; metadati completi; "
        "fattori applicati solo alla biomassa associata."
    )
    c = ws.cell(row=note_row, column=1, value=note)
    c.font = Font(italic=True, size=8, color=SLATE_500)
    c.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[note_row].height = 60


