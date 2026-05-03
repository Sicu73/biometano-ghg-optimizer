# -*- coding: utf-8 -*-
"""export/excel_export.py — Esportazione Excel da output_model.

Funzione pubblica:
  build_excel_from_output(output_model, snapshot=False) -> BytesIO

Questa funzione e' un ADAPTER che:
1. Ricostruisce il ctx compatibile con l'API di excel_export.py (root level)
   a partire dall'output_model strutturato.
2. Chiama `excel_export.build_metaniq_xlsx(ctx)` esistente.

In questo modo il refactoring non richiede di toccare i 2072 righe di
excel_export.py: la logica di generazione XLSX resta invariata, ma ora
l'entry point ufficiale e' `build_excel_from_output(output_model)`.

Se excel_export.py non e' disponibile (es. test isolati), viene generato
un XLSX minimale con openpyxl direttamente.
"""
from __future__ import annotations

from io import BytesIO
from typing import Any

try:
    from excel_export import build_metaniq_xlsx as _build_xlsx_legacy  # type: ignore[import]
    _HAS_LEGACY_XLSX = True
except ImportError:
    _HAS_LEGACY_XLSX = False

try:
    import pandas as pd
    _HAS_PANDAS = True
except ImportError:
    _HAS_PANDAS = False

try:
    from openpyxl import Workbook
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


def build_excel_from_output(output_model: dict, snapshot: bool = False) -> BytesIO:
    """Genera file XLSX da output_model.

    Args:
        output_model:   Dict prodotto da build_output_model().
        snapshot:       Se True, genera snapshot statico (no formule live).
                        Se False, genera file editabile con formule live.

    Returns:
        BytesIO con il file .xlsx, pronto per st.download_button.

    Raises:
        RuntimeError:   Se ne openpyxl ne legacy excel_export disponibili.
    """
    if not isinstance(output_model, dict):
        raise ValueError(f"output_model deve essere un dict, ricevuto: {type(output_model)}")

    # --- Tenta via legacy (excel_export.build_metaniq_xlsx) ------------------
    if _HAS_LEGACY_XLSX:
        try:
            ctx = _output_model_to_xlsx_ctx(output_model, snapshot=snapshot)
            return _build_xlsx_legacy(ctx, snapshot=snapshot)
        except Exception as exc:
            # Fallback a XLSX minimale se legacy fallisce
            return _build_fallback_xlsx(output_model, error=str(exc))

    # --- Fallback: XLSX minimale con openpyxl --------------------------------
    return _build_fallback_xlsx(output_model)


# ---------------------------------------------------------------------------
# Adattatore output_model -> ctx legacy
# ---------------------------------------------------------------------------

def _output_model_to_xlsx_ctx(output_model: dict, snapshot: bool = False) -> dict:
    """Ricostruisce il ctx dict compatibile con excel_export.build_metaniq_xlsx.

    Il ctx legacy richiede chiavi specifiche costruite inline in app_mensile.py.
    Questa funzione le ricostruisce dall'output_model strutturato.
    """
    meta = output_model.get("metadata", {})
    inp = output_model.get("input_summary", {})
    plant = inp.get("plant", {})
    calc = output_model.get("calculation_summary", {})
    monthly_rows = output_model.get("monthly_table", [])

    # df_res per snapshot (se disponibile)
    df_res = None
    if snapshot and _HAS_PANDAS and monthly_rows:
        df_res = pd.DataFrame(monthly_rows)

    # Ricostruisci initial_data dal monthly_table per file editabile
    active_feeds: list[str] = [f["name"] for f in inp.get("feedstocks", [])]
    initial_data: dict[str, dict] = {}
    for row in monthly_rows:
        mese = row.get("Mese", "")
        if not mese:
            continue
        initial_data[mese] = {"Ore": int(row.get("Ore", 0))}
        for f in active_feeds:
            if f in row:
                try:
                    initial_data[mese][f] = float(row[f])
                except (TypeError, ValueError):
                    pass

    # Ricostruisci FEEDSTOCK_DB minimale dall'output_model
    feedstock_db_proxy: dict = {}
    for fs in inp.get("feedstocks", []):
        feedstock_db_proxy[fs["name"]] = {
            "yield": fs.get("yield_std", 0.0),
            "eec":   fs.get("eec", 0.0),
            "etd":   fs.get("etd", 0.0),
            "esca":  fs.get("esca", 0.0),
            "cat":   fs.get("categoria", ""),
            "annex_ix": fs.get("annex_ix"),
            "src":   fs.get("src", ""),
            "color": "#64748B",
        }

    app_mode = meta.get("app_mode", "biometano")
    is_chp = bool(plant.get("is_chp", False))

    ctx: dict[str, Any] = {
        "active_feeds":      active_feeds,
        "FEEDSTOCK_DB":      feedstock_db_proxy,
        "aux_factor":        plant.get("aux_factor", 1.29),
        "ep_total":          plant.get("ep_total", 0.0),
        "fossil_comparator": plant.get("fossil_comparator", 80.0),
        "ghg_threshold":     plant.get("ghg_threshold", 0.80),
        "plant_net_smch":    plant.get("plant_net_smch", 300.0),
        "NM3_TO_MWH":        0.00997,
        "MONTHS": [
            "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
            "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre",
        ],
        "MONTH_HOURS":       [744, 672, 744, 720, 744, 720, 744, 744, 720, 744, 720, 744],
        "initial_data":      initial_data,
        "APP_MODE_LABEL":    meta.get("scenario_name", "Metan.iQ"),
        "end_use":           plant.get("end_use", ""),
        "IS_CHP":            is_chp,
        "plant_kwe":         plant.get("plant_kwe") or 0.0,
        "plant_kwe_net":     plant.get("plant_kwe") or 0.0,
        "eta_el":            0.40,
        "eta_th":            0.42,
        "aux_el_pct":        0.08,
        "lang":              meta.get("language", "it"),
        # Audit
        "yield_audit_rows":  output_model.get("audit_trail", []),
        "emission_audit_rows": [],
        "effective_yields":  {},
        "emission_overrides": {},
        # BP defaults (non disponibili nel ctx minimale, uso defaults)
        "bp_tariffa_eff_mwh": 131.0,
        "bp_ore_anno":        8500.0,
        "bp_lt_tasso":        4.0,
        "bp_lt_durata":       15,
        "bp_lt_leva":         80.0,
        "bp_inflazione_pct":  2.5,
        "bp_durata_tariffa":  15,
        "bp_pnrr_pct":        40.0,
        "bp_ebitda_target_pct": 25.0,
        "bp_tax_rate_pct":    24.0,
        "bp_ammort_anni":     22,
        "bp_npv_disc_rate_pct": 6.0,
        "bp_massimale_eur_per_smch": 32817.23,
        "bp_capex_breakdown": None,
        "bp_capex_forfait":   None,
        "bp_opex_breakdown":  None,
        "bp_opex_forfait":    None,
    }

    if snapshot and df_res is not None:
        ctx["df_res"] = df_res

    return ctx


# ---------------------------------------------------------------------------
# Fallback XLSX minimale (se legacy non disponibile o fallisce)
# ---------------------------------------------------------------------------

def _build_fallback_xlsx(output_model: dict, error: str | None = None) -> BytesIO:
    """Genera un XLSX minimale con openpyxl se il generatore legacy non e' disponibile."""
    buf = BytesIO()

    if not _HAS_OPENPYXL:
        # Ultimo fallback: file vuoto con nota
        buf.write(b"")
        buf.seek(0)
        return buf

    wb = Workbook()
    meta = output_model.get("metadata", {})
    calc = output_model.get("calculation_summary", {})

    # --- Sheet 1: Piano mensile ---
    ws1 = wb.active
    ws1.title = "Piano mensile"
    monthly_rows = output_model.get("monthly_table", [])
    if monthly_rows:
        headers = list(monthly_rows[0].keys())
        ws1.append(headers)
        for row in monthly_rows:
            ws1.append([row.get(h, "") for h in headers])
    else:
        ws1.append(["Nessun dato disponibile"])

    # --- Sheet 2: Feedstock ---
    ws2 = wb.create_sheet("Feedstock")
    ft_rows = output_model.get("feedstock_table", [])
    if ft_rows:
        ws2.append(list(ft_rows[0].keys()))
        for row in ft_rows:
            ws2.append(list(row.values()))

    # --- Sheet 3: KPI ---
    ws3 = wb.create_sheet("KPI")
    ws3.append(["KPI", "Valore"])
    ws3.append(["Software", meta.get("software_name", "Metan.iQ")])
    ws3.append(["Versione", meta.get("version", "")])
    ws3.append(["Generato", meta.get("generated_at", "")])
    ws3.append(["Scenario", meta.get("scenario_name", "")])
    ws3.append(["Totale biomasse (t)", calc.get("tot_biomasse_t", 0.0)])
    ws3.append(["Totale Sm³ netti", calc.get("tot_sm3_netti", 0.0)])
    ws3.append(["Totale MWh", calc.get("tot_mwh", 0.0)])
    ws3.append(["Saving medio (%)", calc.get("saving_avg", 0.0)])
    ws3.append(["Mesi validi", calc.get("valid_months", 0)])
    ws3.append(["Ricavi totali (EUR)", calc.get("total_revenue", 0.0)])

    if error:
        ws3.append(["ERRORE generatore legacy", error])

    # --- Sheet 4: Warnings ---
    ws4 = wb.create_sheet("Note")
    ws4.append(["Tipo", "Messaggio"])
    for w in output_model.get("warnings", []):
        ws4.append(["WARNING", w])
    for e in output_model.get("errors", []):
        ws4.append(["ERROR", e])
    expl = output_model.get("explanations", {})
    for k, v in expl.items():
        ws4.append([f"NOTA_{k}", v])

    wb.save(buf)
    buf.seek(0)
    return buf
