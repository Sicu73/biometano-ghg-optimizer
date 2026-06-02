# -*- coding: utf-8 -*-
# Copyright (c) 2026 Carlo Sicurini. All Rights Reserved.
# Metan.iQ - Biometano GHG Optimizer (DM 2022 / RED III)
# Proprietary and confidential. See LICENSE for terms.
# Commercial licensing: carlo.sicurini@gmail.com
"""export/daily_excel.py — Esportazione Excel giornaliero (multi-foglio, stile Metan.iQ).

Genera un .xlsx professionale con:
    - Riepilogo Mensile (titolo, periodo, esito ufficiale, KPI principali)
    - Operatività Giornaliera (tabella daily con formattazione IT)
    - Confronto Standard vs Analisi (se attivi override BMT/EF)
    - Override Attivi (BMT lab + EF su misura)
    - Anagrafica & Parametri (audit trail formattato)
    - Vincoli (constraints_status)

Style: brand Metan.iQ — navy `#1E3A5F`, accent orange `#F59E0B`, success green
`#0F9D58`, danger red `#DB4437`. Font Arial. Formattazione numeri locale-agnostica
(Excel renderizza secondo locale IT). Freeze panes, bordi sottili, righe alternate.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any

import pandas as pd

# ============================================================================
# BRAND PALETTE (Metan.iQ)
# ============================================================================
_NAVY = "1E3A5F"           # header principali
_NAVY_DARK = "0F1F33"      # titolo H1
_ORANGE = "F59E0B"         # accent / sottotitoli
_TINT = "E8F0F7"           # banda alternata righe
_TINT_DARK = "C9DCEC"      # banda header tabelle
_GREEN = "0F9D58"          # OK / compliant
_RED = "DB4437"            # NOT compliant
_GREY_BORDER = "B8C2CC"
_WHITE = "FFFFFF"

# ============================================================================
# FORMATI NUMERO (locale-agnostici — Excel applica separatore secondo locale)
# ============================================================================
_FMT_INT = "#,##0"
_FMT_DEC1 = "#,##0.0"
_FMT_DEC2 = "#,##0.00"
_FMT_DEC3 = "#,##0.000"
_FMT_PCT2 = "0.00%"
_FMT_PCT_PT = '0.00" pt"'   # delta in punti percentuali
_FMT_DATE = "dd/mm/yyyy"


def _it_num(value: float, decimals: int = 2) -> str:
    """Formatta numero in stile italiano (1.234,56) per stringhe."""
    try:
        s = f"{float(value):,.{decimals}f}"
        return s.replace(",", "X").replace(".", ",").replace("X", ".")
    except (TypeError, ValueError):
        return str(value)


# ============================================================================
# BUILDER PRINCIPALE
# ============================================================================
def build_daily_excel(
    daily_df: pd.DataFrame,
    monthly_kpis: dict[str, Any],
    audit_trail: dict[str, Any] | None = None,
    *,
    kpis_standard: dict[str, Any] | None = None,
    overrides_info: dict[str, Any] | None = None,
    plant_info: dict[str, Any] | None = None,
) -> bytes:
    """Crea un .xlsx Metan.iQ con riepilogo + dettagli giornalieri + override.

    Args:
        daily_df: tabella giornaliera (Data, ore_funz, biomasse, sm³ lordi/netti,
            MWh, saving %, esito).
        monthly_kpis: dict KPI mensili (sm3_netti, mwh, saving_pct, threshold,
            compliant, biomass_total_t, feedstock_totals, constraints_status, ...).
        audit_trail: dict {label: valore} con parametri/anagrafica/configurazione.
        kpis_standard: opzionale, KPI ricalcolati senza override (vista Standard).
            Se presente, abilita foglio "Confronto Standard vs Analisi".
        overrides_info: opzionale, {"bmt": {feed: yield}, "ef": {feed: ef_dict}}.
            Se non vuoto, abilita foglio "Override Attivi".
        plant_info: opzionale, metadati periodo (year, month, regime, threshold,
            company, plant_name, ...) per popolare il foglio di riepilogo.

    Returns:
        Bytes del file .xlsx pronto per `st.download_button`.
    """
    audit_trail = audit_trail or {}
    overrides_info = overrides_info or {}
    plant_info = plant_info or {}

    # Fallback grazioso se openpyxl non disponibile
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Alignment, Border, Font, PatternFill, Side,
        )
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback: CSV-like (mai dovrebbe scattare — openpyxl è in requirements)
        return daily_df.to_csv(sep=";", index=False).encode("utf-8")

    # ------------------------------------------------------------------
    # STILI RIUTILIZZABILI
    # ------------------------------------------------------------------
    _font_base = Font(name="Arial", size=10)
    _font_bold = Font(name="Arial", size=10, bold=True)
    _font_h1 = Font(name="Arial", size=18, bold=True, color=_WHITE)
    _font_h2 = Font(name="Arial", size=12, bold=True, color=_WHITE)
    _font_h3 = Font(name="Arial", size=11, bold=True, color=_NAVY_DARK)
    _font_kpi_label = Font(name="Arial", size=9, bold=True, color="595959")
    _font_kpi_value = Font(name="Arial", size=18, bold=True, color=_NAVY_DARK)
    _font_caption = Font(name="Arial", size=8, italic=True, color="6E6E6E")

    _fill_navy = PatternFill("solid", start_color=_NAVY)
    _fill_navy_dark = PatternFill("solid", start_color=_NAVY_DARK)
    _fill_orange = PatternFill("solid", start_color=_ORANGE)
    _fill_tint = PatternFill("solid", start_color=_TINT)
    _fill_tint_dark = PatternFill("solid", start_color=_TINT_DARK)
    _fill_green = PatternFill("solid", start_color=_GREEN)
    _fill_red = PatternFill("solid", start_color=_RED)
    _fill_white = PatternFill("solid", start_color=_WHITE)

    _side = Side(style="thin", color=_GREY_BORDER)
    _border = Border(left=_side, right=_side, top=_side, bottom=_side)

    _center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    _right = Alignment(horizontal="right", vertical="center")

    # ------------------------------------------------------------------
    # WORKBOOK
    # ------------------------------------------------------------------
    wb = Workbook()
    # Rimuovo il foglio default — lo sostituirò con "Riepilogo"
    wb.remove(wb.active)

    # ==================================================================
    # FOGLIO 1 — RIEPILOGO MENSILE
    # ==================================================================
    ws = wb.create_sheet("Riepilogo Mensile")
    ws.sheet_view.showGridLines = False

    # H1 — Banner brand
    ws.merge_cells("A1:F2")
    h1 = ws["A1"]
    h1.value = "Metan.iQ — Operatività Giornaliera"
    h1.font = _font_h1
    h1.fill = _fill_navy_dark
    h1.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # H2 — Periodo + impianto
    ws.merge_cells("A3:F3")
    _year = plant_info.get("year", "")
    _month = plant_info.get("month", "")
    _periodo = f"Periodo: {int(_month):02d}/{int(_year)}" if _year and _month else "Periodo: —"
    _company = plant_info.get("company") or audit_trail.get("Nome azienda", "—")
    _plant = plant_info.get("plant_name") or audit_trail.get("Nome impianto", "—")
    h2 = ws["A3"]
    h2.value = f"{_periodo}  ·  Azienda: {_company}  ·  Impianto: {_plant}"
    h2.font = _font_h2
    h2.fill = _fill_navy
    h2.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Caption esportazione
    ws.merge_cells("A4:F4")
    c = ws["A4"]
    c.value = (
        f"Esportato il {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  "
        f"Regime: {plant_info.get('regime') or audit_trail.get('Regime applicato', '—')}  ·  "
        f"Soglia normativa: {plant_info.get('threshold_pct', monthly_kpis.get('threshold', 0)):.2f}%"
    )
    c.font = _font_caption
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Banner ESITO UFFICIALE
    ws.row_dimensions[6].height = 38
    ws.merge_cells("A6:F6")
    e = ws["A6"]
    _compliant = bool(monthly_kpis.get("compliant", False))
    _saving = float(monthly_kpis.get("saving_pct", 0.0))
    _thr = float(monthly_kpis.get("threshold", 0.0))
    if _compliant:
        e.value = f"✅  ESITO UFFICIALE — DM 2022 (RED III) — SOSTENIBILE   ·   Saving GHG: {_it_num(_saving, 2)}%  (soglia {_it_num(_thr, 2)}%)"
        e.fill = _fill_green
    else:
        e.value = f"⚠  ESITO UFFICIALE — DM 2022 (RED III) — NON CONFORME   ·   Saving GHG: {_it_num(_saving, 2)}%  (soglia {_it_num(_thr, 2)}%)"
        e.fill = _fill_red
    e.font = Font(name="Arial", size=14, bold=True, color=_WHITE)
    e.alignment = Alignment(horizontal="center", vertical="center")

    # KPI cards (4 in riga)
    ws.row_dimensions[8].height = 16
    ws.row_dimensions[9].height = 32

    _kpi_cards = [
        ("🌾 Biomassa mese",     monthly_kpis.get("biomass_total_t", 0.0), _FMT_DEC1, "t"),
        ("⚡ MWh netti mese",   monthly_kpis.get("mwh", 0.0),              _FMT_DEC1, "MWh"),
        ("🟢 Sm³ netti mese",   monthly_kpis.get("sm3_netti", 0.0),        _FMT_INT,  "Sm³"),
        ("🌱 Saving GHG",       _saving / 100.0,                            _FMT_PCT2, ""),
    ]
    _cols_pairs = [("A", "B"), ("C", "C"), ("D", "D"), ("E", "F")]
    # Più semplice: card su 6 colonne, una ogni 1-2 colonne
    _start_cols = ["A", "B", "C", "D"]  # colonne separate per chiarezza
    # Rifaccio layout: 4 colonne × 2 righe (label sopra, value sotto)
    for i, (label, value, fmt, unit) in enumerate(_kpi_cards):
        col_letter = chr(ord("A") + i)  # A, B, C, D
        # Estendo card su 1 colonna ciascuna (4 colonne usate, E-F per padding)
        lbl_cell = ws[f"{col_letter}8"]
        val_cell = ws[f"{col_letter}9"]
        lbl_cell.value = label + (f"  ({unit})" if unit else "")
        lbl_cell.font = _font_kpi_label
        lbl_cell.fill = _fill_tint
        lbl_cell.alignment = _center
        lbl_cell.border = _border
        val_cell.value = value
        val_cell.font = _font_kpi_value
        val_cell.fill = _fill_white
        val_cell.alignment = _center
        val_cell.border = _border
        val_cell.number_format = fmt

    # Tabella riepilogo feedstock (se presente)
    _feed_totals = monthly_kpis.get("feedstock_totals") or {}
    _row = 12
    if _feed_totals:
        ws.merge_cells(f"A{_row}:F{_row}")
        h = ws.cell(row=_row, column=1, value="🌾 Biomasse — totali mensili (t)")
        h.font = _font_h3
        h.alignment = _left
        _row += 1
        # Header tabella
        ws.cell(row=_row, column=1, value="Biomassa").font = _font_bold
        ws.cell(row=_row, column=2, value="Tonnellate").font = _font_bold
        for c_ in (1, 2):
            cell = ws.cell(row=_row, column=c_)
            cell.fill = _fill_tint_dark
            cell.alignment = _center
            cell.border = _border
        _row += 1
        for fname, qty in _feed_totals.items():
            ws.cell(row=_row, column=1, value=str(fname)).alignment = _left
            ws.cell(row=_row, column=2, value=float(qty)).number_format = _FMT_DEC1
            for c_ in (1, 2):
                cell = ws.cell(row=_row, column=c_)
                cell.font = _font_base
                cell.border = _border
            _row += 1
        # Totale
        ws.cell(row=_row, column=1, value="∑ TOTALE").font = _font_bold
        ws.cell(row=_row, column=2, value=f"=SUM(B{_row - len(_feed_totals)}:B{_row - 1})").font = _font_bold
        ws.cell(row=_row, column=2).number_format = _FMT_DEC1
        for c_ in (1, 2):
            cell = ws.cell(row=_row, column=c_)
            cell.fill = _fill_tint
            cell.border = _border

    # Larghezze colonne
    for col_letter, width in zip("ABCDEF", [26, 22, 22, 22, 16, 16]):
        ws.column_dimensions[col_letter].width = width

    # ==================================================================
    # FOGLIO 2 — OPERATIVITÀ GIORNALIERA
    # ==================================================================
    ws2 = wb.create_sheet("Operatività Giornaliera")
    ws2.sheet_view.showGridLines = False

    df_day = daily_df.copy()
    if "Data" in df_day.columns:
        df_day["Data"] = pd.to_datetime(df_day["Data"], errors="coerce")

    # Titolo
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(df_day.columns), 1))
    t = ws2.cell(row=1, column=1, value="📆 Operatività Giornaliera — Tabella ore funzionamento + biomasse")
    t.font = _font_h2
    t.fill = _fill_navy
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws2.row_dimensions[1].height = 26

    # Header colonne (riga 3)
    headers = list(df_day.columns)
    for j, col in enumerate(headers, start=1):
        cell = ws2.cell(row=3, column=j, value=str(col))
        cell.font = Font(name="Arial", size=10, bold=True, color=_WHITE)
        cell.fill = _fill_navy
        cell.alignment = _center
        cell.border = _border
    ws2.row_dimensions[3].height = 30

    # Righe dati (riga 4+)
    _date_cols = {i for i, c in enumerate(headers, start=1) if "data" in str(c).lower()}
    _pct_cols = {i for i, c in enumerate(headers, start=1)
                 if "%" in str(c) or "saving" in str(c).lower()}
    _int_cols = {i for i, c in enumerate(headers, start=1)
                 if "sm³" in str(c).lower() or "sm3" in str(c).lower() or "smc" in str(c).lower()}

    for r_idx, row in enumerate(df_day.itertuples(index=False, name=None), start=4):
        for j, val in enumerate(row, start=1):
            cell = ws2.cell(row=r_idx, column=j, value=val)
            cell.font = _font_base
            cell.border = _border
            if j in _date_cols:
                cell.number_format = _FMT_DATE
                cell.alignment = _center
            elif j in _pct_cols:
                cell.number_format = _FMT_DEC2
                cell.alignment = _right
            elif j in _int_cols:
                cell.number_format = _FMT_INT
                cell.alignment = _right
            elif isinstance(val, (int, float)):
                cell.number_format = _FMT_DEC2
                cell.alignment = _right
            else:
                cell.alignment = _left
            # Banda alternata
            if (r_idx - 4) % 2 == 1:
                cell.fill = _fill_tint

    # Freeze su header
    ws2.freeze_panes = "A4"

    # Larghezze colonne (auto-fit basato su header + sample)
    for j, col in enumerate(headers, start=1):
        max_len = max(len(str(col)), 10)
        # campione primi 30 valori
        try:
            for sample in df_day.iloc[:30, j - 1]:
                max_len = max(max_len, len(str(sample)))
        except Exception:
            pass
        ws2.column_dimensions[get_column_letter(j)].width = min(max_len + 3, 28)

    # Conditional formatting su colonna saving %
    for j in _pct_cols:
        col_l = get_column_letter(j)
        rng = f"{col_l}4:{col_l}{3 + len(df_day)}"
        # Verde se >= soglia
        ws2.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="greaterThanOrEqual", formula=[f"{_thr}"],
                fill=PatternFill("solid", start_color="D9F2D9"),
                font=Font(name="Arial", size=10, color="0B5C0B", bold=True),
            ),
        )
        ws2.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="lessThan", formula=[f"{_thr}"],
                fill=PatternFill("solid", start_color="FBE0DC"),
                font=Font(name="Arial", size=10, color="8B1A0E", bold=True),
            ),
        )

    # ==================================================================
    # FOGLIO 3 — CONFRONTO STANDARD vs ANALISI (opzionale)
    # ==================================================================
    if kpis_standard:
        ws3 = wb.create_sheet("Confronto Standard vs Analisi")
        ws3.sheet_view.showGridLines = False

        ws3.merge_cells("A1:D1")
        t = ws3["A1"]
        t.value = "📊 Confronto Standard vs Analisi (BMT lab + EF su misura)"
        t.font = _font_h2
        t.fill = _fill_navy
        t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws3.row_dimensions[1].height = 26

        # Caption
        ws3.merge_cells("A2:D2")
        cap = ws3["A2"]
        cap.value = (
            "Standard = UNI-TS / RED III tabellare. Analisi = override BMT laboratorio "
            "e/o Fattori Emissivi da relazione tecnica."
        )
        cap.font = _font_caption
        cap.alignment = _left

        # Header
        for j, lbl in enumerate(["KPI", "Standard", "Analisi", "Δ %"], start=1):
            cell = ws3.cell(row=4, column=j, value=lbl)
            cell.font = Font(name="Arial", size=10, bold=True, color=_WHITE)
            cell.fill = _fill_navy
            cell.alignment = _center
            cell.border = _border
        ws3.row_dimensions[4].height = 24

        # Righe KPI confronto
        _comp_rows = [
            ("Biomasse (t/mese)",   "biomass_total_t", _FMT_DEC1, False),
            ("Sm³ netti (mese)",    "sm3_netti",       _FMT_INT,  False),
            ("MWh netti (mese)",    "mwh",             _FMT_DEC1, False),
            ("Saving GHG (%)",      "saving_pct",      _FMT_DEC2, True),
        ]
        for i, (label, key, fmt, is_pct_pt) in enumerate(_comp_rows, start=5):
            std_v = float(kpis_standard.get(key, 0.0))
            ana_v = float(monthly_kpis.get(key, 0.0))
            if is_pct_pt:
                delta = ana_v - std_v
                delta_fmt = _FMT_PCT_PT
            else:
                delta = ((ana_v - std_v) / std_v * 100.0) if std_v else 0.0
                delta_fmt = '+0.00"%";-0.00"%";"—"'
            ws3.cell(row=i, column=1, value=label).alignment = _left
            ws3.cell(row=i, column=2, value=std_v).number_format = fmt
            ws3.cell(row=i, column=3, value=ana_v).number_format = fmt
            ws3.cell(row=i, column=4, value=delta).number_format = delta_fmt
            for c_ in (1, 2, 3, 4):
                cell = ws3.cell(row=i, column=c_)
                cell.font = _font_base
                cell.border = _border
                if c_ > 1:
                    cell.alignment = _right
                if i % 2 == 1:
                    cell.fill = _fill_tint

        # Compliance comparison banner
        std_compliant = bool(kpis_standard.get("compliant", False))
        ana_compliant = bool(monthly_kpis.get("compliant", False))
        if std_compliant != ana_compliant:
            ws3.row_dimensions[11].height = 30
            ws3.merge_cells("A11:D11")
            b = ws3["A11"]
            if ana_compliant and not std_compliant:
                b.value = "✅ Gli override (BMT/EF) rendono il mese SOSTENIBILE. Con i valori Standard NON saresti conforme."
                b.fill = _fill_green
            else:
                b.value = "⚠ Con i valori Standard il mese sarebbe SOSTENIBILE, ma gli override abbassano la performance sotto soglia."
                b.fill = _fill_red
            b.font = Font(name="Arial", size=10, bold=True, color=_WHITE)
            b.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_letter, width in zip("ABCD", [32, 16, 16, 14]):
            ws3.column_dimensions[col_letter].width = width

    # ==================================================================
    # FOGLIO 4 — OVERRIDE ATTIVI (opzionale)
    # ==================================================================
    _bmt = overrides_info.get("bmt") or {}
    _ef = overrides_info.get("ef") or {}
    if _bmt or _ef:
        ws4 = wb.create_sheet("Override Attivi")
        ws4.sheet_view.showGridLines = False

        ws4.merge_cells("A1:E1")
        t = ws4["A1"]
        t.value = "🧪 Override Attivi — BMT lab & Fattori Emissivi su misura"
        t.font = _font_h2
        t.fill = _fill_navy
        t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws4.row_dimensions[1].height = 26

        _row = 3
        if _bmt:
            ws4.merge_cells(f"A{_row}:E{_row}")
            h = ws4.cell(row=_row, column=1, value=f"🧪 BMT Override ({len(_bmt)} biomasse)")
            h.font = _font_h3
            h.alignment = _left
            _row += 1
            for j, lbl in enumerate(["Biomassa", "Resa BMT (Sm³/t)"], start=1):
                cell = ws4.cell(row=_row, column=j, value=lbl)
                cell.font = _font_bold
                cell.fill = _fill_tint_dark
                cell.alignment = _center
                cell.border = _border
            _row += 1
            for fname, yld in _bmt.items():
                ws4.cell(row=_row, column=1, value=str(fname)).alignment = _left
                ws4.cell(row=_row, column=2, value=float(yld)).number_format = _FMT_DEC2
                for c_ in (1, 2):
                    cell = ws4.cell(row=_row, column=c_)
                    cell.font = _font_base
                    cell.border = _border
                if c_ == 2:
                    cell.alignment = _right
                _row += 1
            _row += 1  # gap

        if _ef:
            ws4.merge_cells(f"A{_row}:E{_row}")
            h = ws4.cell(row=_row, column=1, value=f"🧬 Fattori Emissivi Override ({len(_ef)} biomasse)")
            h.font = _font_h3
            h.alignment = _left
            _row += 1
            for j, lbl in enumerate(
                ["Biomassa", "e_ec", "e_l", "e_p", "e_td"], start=1
            ):
                cell = ws4.cell(row=_row, column=j, value=lbl)
                cell.font = _font_bold
                cell.fill = _fill_tint_dark
                cell.alignment = _center
                cell.border = _border
            _row += 1
            for fname, ef_dict in _ef.items():
                if not isinstance(ef_dict, dict):
                    continue
                ws4.cell(row=_row, column=1, value=str(fname)).alignment = _left
                ws4.cell(row=_row, column=2, value=float(ef_dict.get("e_ec", 0))).number_format = _FMT_DEC3
                ws4.cell(row=_row, column=3, value=float(ef_dict.get("e_l", 0))).number_format = _FMT_DEC3
                ws4.cell(row=_row, column=4, value=float(ef_dict.get("e_p", 0))).number_format = _FMT_DEC3
                ws4.cell(row=_row, column=5, value=float(ef_dict.get("e_td", 0))).number_format = _FMT_DEC3
                for c_ in range(1, 6):
                    cell = ws4.cell(row=_row, column=c_)
                    cell.font = _font_base
                    cell.border = _border
                    if c_ > 1:
                        cell.alignment = _right
                _row += 1

        for col_letter, width in zip("ABCDE", [34, 14, 14, 14, 14]):
            ws4.column_dimensions[col_letter].width = width

    # ==================================================================
    # FOGLIO 5 — ANAGRAFICA & PARAMETRI
    # ==================================================================
    ws5 = wb.create_sheet("Anagrafica & Parametri")
    ws5.sheet_view.showGridLines = False

    ws5.merge_cells("A1:C1")
    t = ws5["A1"]
    t.value = "📋 Anagrafica simulazione & Parametri di calcolo"
    t.font = _font_h2
    t.fill = _fill_navy
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws5.row_dimensions[1].height = 26

    # Header
    for j, lbl in enumerate(["Voce", "Valore"], start=1):
        cell = ws5.cell(row=3, column=j, value=lbl)
        cell.font = _font_bold
        cell.fill = _fill_tint_dark
        cell.alignment = _center
        cell.border = _border
    ws5.row_dimensions[3].height = 22

    _row = 4
    for k, v in audit_trail.items():
        if isinstance(v, (list, tuple)):
            v = " | ".join(str(x) for x in v)
        ws5.cell(row=_row, column=1, value=str(k)).alignment = _left
        ws5.cell(row=_row, column=2, value=str(v) if v is not None else "—").alignment = _left
        for c_ in (1, 2):
            cell = ws5.cell(row=_row, column=c_)
            cell.font = _font_base
            cell.border = _border
            if _row % 2 == 1:
                cell.fill = _fill_tint
        _row += 1

    ws5.column_dimensions["A"].width = 42
    ws5.column_dimensions["B"].width = 58
    ws5.freeze_panes = "A4"

    # ==================================================================
    # FOGLIO 6 — VINCOLI
    # ==================================================================
    constraints = monthly_kpis.get("constraints_status") or []
    if constraints:
        ws6 = wb.create_sheet("Vincoli")
        ws6.sheet_view.showGridLines = False

        df_con = pd.DataFrame(constraints)
        if df_con.empty:
            df_con = pd.DataFrame(columns=["name", "ok", "value", "limit", "msg"])

        ws6.merge_cells(start_row=1, start_column=1,
                        end_row=1, end_column=max(len(df_con.columns), 1))
        t = ws6.cell(row=1, column=1, value="✅ Vincoli normativi — verifica per regime")
        t.font = _font_h2
        t.fill = _fill_navy
        t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws6.row_dimensions[1].height = 26

        headers_c = list(df_con.columns)
        for j, col in enumerate(headers_c, start=1):
            cell = ws6.cell(row=3, column=j, value=str(col))
            cell.font = Font(name="Arial", size=10, bold=True, color=_WHITE)
            cell.fill = _fill_navy
            cell.alignment = _center
            cell.border = _border
        ws6.row_dimensions[3].height = 22

        _ok_col_idx = None
        for j, col in enumerate(headers_c, start=1):
            if str(col).lower() in ("ok", "compliant", "valid"):
                _ok_col_idx = j

        for r_idx, row in enumerate(df_con.itertuples(index=False, name=None), start=4):
            for j, val in enumerate(row, start=1):
                cell = ws6.cell(row=r_idx, column=j, value=val)
                cell.font = _font_base
                cell.border = _border
                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    cell.number_format = _FMT_DEC2
                    cell.alignment = _right
                else:
                    cell.alignment = _left
                if j == _ok_col_idx and val is not None:
                    if bool(val):
                        cell.fill = _fill_green
                        cell.font = Font(name="Arial", size=10, bold=True, color=_WHITE)
                        cell.alignment = _center
                        cell.value = "✓ OK"
                    else:
                        cell.fill = _fill_red
                        cell.font = Font(name="Arial", size=10, bold=True, color=_WHITE)
                        cell.alignment = _center
                        cell.value = "✗ KO"
                elif (r_idx - 4) % 2 == 1:
                    cell.fill = _fill_tint

        for j, col in enumerate(headers_c, start=1):
            max_len = max(len(str(col)), 12)
            try:
                for sample in df_con.iloc[:30, j - 1]:
                    max_len = max(max_len, len(str(sample)))
            except Exception:
                pass
            ws6.column_dimensions[get_column_letter(j)].width = min(max_len + 3, 40)
        ws6.freeze_panes = "A4"

    # ------------------------------------------------------------------
    # SAVE → bytes
    # ------------------------------------------------------------------
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = ["build_daily_excel"]
