# -*- coding: utf-8 -*-
"""Automatic bilingual runtime patch for Metan.iQ.

Python imports this file automatically at process startup when the project root
is on sys.path. The patch keeps the calculation engine unchanged and translates
UI labels plus CSV/XLSX/PDF exports when st.session_state['lang'] == 'en'.
"""
from __future__ import annotations

from io import BytesIO

try:
    from i18n_runtime import LANG_EN, tr, translate_columns, export_filename
except Exception:  # pragma: no cover
    LANG_EN = "en"
    def tr(text, lang="it"):
        return text
    def translate_columns(df, lang="it"):
        return df
    def export_filename(base, lang, ext):
        return f"{base}.{ext.lstrip('.')}"


REPLACEMENTS_EN = {
    "Piano mensile": "Monthly planner",
    "Sintesi annuale": "Annual summary",
    "Database feedstock": "Feedstock database",
    "Scarica CSV": "Download CSV",
    "Scarica Excel": "Download Excel",
    "Scarica PDF": "Download PDF",
    "Excel modificabile": "Editable Excel",
    "Report PDF": "PDF report",
    "Tabella risultati": "Results table",
    "Risultati mensili": "Monthly results",
    "Parametri impianto": "Plant parameters",
    "Destinazione": "End use",
    "Modalità": "Mode",
    "Generato il": "Generated on",
    "Validità": "Validity",
    "Valido": "Valid",
    "Non valido": "Invalid",
    "Conforme": "Compliant",
    "Non conforme": "Not compliant",
    "Risparmio GHG": "GHG saving",
    "Emissioni": "Emissions",
    "Biomassa": "Feedstock",
    "Biomasse": "Feedstocks",
    "Produzione": "Production",
    "Ricavi": "Revenue",
    "Ricavo": "Revenue",
    "Costi": "Costs",
    "Costo": "Cost",
    "Margine": "Margin",
    "Energia": "Energy",
    "Elettricità": "Electricity",
    "Calore": "Heat",
    "Mese": "Month",
    "Ore": "Hours",
    "Anno": "Year",
    "Totale": "Total",
    "Valore": "Value",
    "Parametro": "Parameter",
    "Note": "Notes",
    "Soglia": "Threshold",
    "Esito": "Outcome",
    "Unità": "Unit",
    "Liquame suino": "Pig slurry",
    "Pollina ovaiole": "Layer manure",
    "Trinciato di mais": "Maize silage",
    "Trinciato di sorgo": "Sorghum silage",
    "Gennaio": "January", "Febbraio": "February", "Marzo": "March",
    "Aprile": "April", "Maggio": "May", "Giugno": "June",
    "Luglio": "July", "Agosto": "August", "Settembre": "September",
    "Ottobre": "October", "Novembre": "November", "Dicembre": "December",
}


def _active_lang():
    try:
        import streamlit as st
        return st.session_state.get("lang", "it")
    except Exception:
        return "it"


def _t(value):
    if _active_lang() != LANG_EN or not isinstance(value, str):
        return value
    out = tr(value, LANG_EN)
    if out != value:
        return out
    for it, en in sorted(REPLACEMENTS_EN.items(), key=lambda kv: len(kv[0]), reverse=True):
        out = out.replace(it, en)
    return out


def _translate_df(obj):
    try:
        import pandas as pd
        if isinstance(obj, pd.DataFrame) and _active_lang() == LANG_EN:
            return translate_columns(obj, LANG_EN)
    except Exception:
        pass
    return obj


def _to_bytes(data):
    if isinstance(data, BytesIO):
        pos = data.tell()
        data.seek(0)
        raw = data.read()
        data.seek(pos)
        return raw
    if isinstance(data, bytes):
        return data
    if isinstance(data, str):
        return data.encode("utf-8")
    return None


def _translate_csv_bytes(raw):
    try:
        text = raw.decode("utf-8-sig")
    except Exception:
        return raw
    return _t(text).encode("utf-8-sig")


def _translate_xlsx_bytes(raw):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(raw))
        for ws in wb.worksheets:
            ws.title = _t(ws.title)[:31]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and not cell.value.startswith("="):
                        cell.value = _t(cell.value)
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out
    except Exception:
        return BytesIO(raw)


def _patch_streamlit():
    try:
        import streamlit as st
    except Exception:
        return
    if getattr(st, "_metaniq_i18n_patched", False):
        return

    def wrap_text_func(name):
        original = getattr(st, name, None)
        if original is None:
            return
        def wrapped(*args, **kwargs):
            args = list(args)
            if args:
                args[0] = _t(args[0])
            for key in ("label", "body", "help", "placeholder"):
                if key in kwargs:
                    kwargs[key] = _t(kwargs[key])
            return original(*args, **kwargs)
        setattr(st, name, wrapped)

    for name in ["title", "header", "subheader", "caption", "markdown", "write", "button", "radio", "selectbox", "multiselect", "number_input", "text_input", "slider", "expander"]:
        wrap_text_func(name)

    original_tabs = getattr(st, "tabs", None)
    if original_tabs is not None:
        def tabs_wrapped(tabs, *args, **kwargs):
            if _active_lang() == LANG_EN:
                tabs = [_t(x) for x in tabs]
            return original_tabs(tabs, *args, **kwargs)
        st.tabs = tabs_wrapped

    original_dataframe = getattr(st, "dataframe", None)
    if original_dataframe is not None:
        def dataframe_wrapped(data=None, *args, **kwargs):
            return original_dataframe(_translate_df(data), *args, **kwargs)
        st.dataframe = dataframe_wrapped

    original_table = getattr(st, "table", None)
    if original_table is not None:
        def table_wrapped(data=None, *args, **kwargs):
            return original_table(_translate_df(data), *args, **kwargs)
        st.table = table_wrapped

    original_download = getattr(st, "download_button", None)
    if original_download is not None:
        def download_wrapped(*args, **kwargs):
            if args:
                args = list(args)
                args[0] = _t(args[0])
            if "label" in kwargs:
                kwargs["label"] = _t(kwargs["label"])
            fname = kwargs.get("file_name") or ""
            if _active_lang() == LANG_EN and fname:
                lower = fname.lower()
                raw = _to_bytes(kwargs.get("data"))
                if lower.endswith(".csv") and raw is not None:
                    kwargs["data"] = _translate_csv_bytes(raw)
                    kwargs["file_name"] = export_filename(fname.rsplit(".", 1)[0], LANG_EN, "csv")
                elif lower.endswith(".xlsx") and raw is not None:
                    kwargs["data"] = _translate_xlsx_bytes(raw)
                    kwargs["file_name"] = export_filename(fname.rsplit(".", 1)[0], LANG_EN, "xlsx")
                elif lower.endswith(".pdf"):
                    kwargs["file_name"] = export_filename(fname.rsplit(".", 1)[0], LANG_EN, "pdf")
            return original_download(*args, **kwargs)
        st.download_button = download_wrapped

    st._metaniq_i18n_patched = True


def _build_basic_pdf_en(ctx):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    story = [Paragraph("Metan.iQ - Monthly Planning Report", styles["Title"]), Spacer(1, 12)]
    rows = [["Parameter", "Value"]]
    for key in ["APP_MODE_LABEL", "end_use", "tot_biomasse_t", "tot_sm3_netti", "tot_mwh_netti", "saving_avg", "tot_revenue", "valid_months"]:
        if key in ctx:
            rows.append([_t(str(key)), str(ctx.get(key))])
    table = Table(rows, repeatRows=1)
    table.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,0), colors.HexColor("#0F172A")), ("TEXTCOLOR", (0,0), (-1,0), colors.white), ("GRID", (0,0), (-1,-1), 0.25, colors.grey)]))
    story.append(table)
    df = ctx.get("df_res")
    try:
        df = translate_columns(df, LANG_EN)
        if df is not None:
            small = [list(map(str, df.columns[:8]))]
            for _, row in df.head(12).iterrows():
                small.append([str(x) for x in list(row.values[:8])])
            story.extend([Spacer(1, 16), Paragraph("Monthly table", styles["Heading2"]), Table(small, repeatRows=1)])
    except Exception:
        pass
    doc.build(story)
    buf.seek(0)
    return buf


def _patch_exports():
    try:
        import excel_export
        if not getattr(excel_export, "_metaniq_i18n_patched", False):
            original_xlsx = excel_export.build_metaniq_xlsx
            def xlsx_wrapped(ctx, *args, **kwargs):
                out = original_xlsx(ctx, *args, **kwargs)
                lang = ctx.get("lang") if isinstance(ctx, dict) else None
                if lang == LANG_EN or _active_lang() == LANG_EN:
                    raw = _to_bytes(out)
                    if raw is not None:
                        return _translate_xlsx_bytes(raw)
                return out
            excel_export.build_metaniq_xlsx = xlsx_wrapped
            excel_export._metaniq_i18n_patched = True
    except Exception:
        pass

    try:
        import report_pdf
        if not getattr(report_pdf, "_metaniq_i18n_patched", False):
            original_pdf = report_pdf.build_metaniq_pdf
            def pdf_wrapped(ctx, *args, **kwargs):
                lang = ctx.get("lang") if isinstance(ctx, dict) else None
                if lang == LANG_EN or _active_lang() == LANG_EN:
                    try:
                        return _build_basic_pdf_en(ctx)
                    except Exception:
                        return original_pdf(ctx, *args, **kwargs)
                return original_pdf(ctx, *args, **kwargs)
            report_pdf.build_metaniq_pdf = pdf_wrapped
            report_pdf._metaniq_i18n_patched = True
    except Exception:
        pass


_patch_streamlit()
_patch_exports()
