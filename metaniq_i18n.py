"""Metan.iQ — layer i18n IT/EN.

Obiettivo: fornire un layer di traduzione centralizzato per UI Streamlit,
tabelle, CSV ed Excel senza duplicare tutta l'app.

Limite dichiarato:
- il PDF ReportLab richiede una migrazione dedicata dei template testuali;
- questo modulo traduce automaticamente UI/CSV/XLSX dove i testi sono esposti
  come stringhe, ma non può garantire la traduzione di testo già rasterizzato o
  incorporato in un PDF già generato.
"""
from __future__ import annotations

from io import BytesIO
from typing import Any

try:
    import pandas as pd
except Exception:  # pragma: no cover
    pd = None

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

LANG_IT = "it"
LANG_EN = "en"

LANG_LABELS = {
    LANG_IT: "Italiano",
    LANG_EN: "English",
}

# Dizionario pragmatico: termini e frasi ricorrenti nell'app Metan.iQ.
# Deve restare volutamente leggibile e manutenibile.
IT_TO_EN: dict[str, str] = {
    # Generale
    "Lingua": "Language",
    "Italiano": "Italian",
    "Inglese": "English",
    "Configurazione": "Configuration",
    "Parametri": "Parameters",
    "Parametri impianto": "Plant parameters",
    "Normativa": "Regulation",
    "Aggiornamenti": "Updates",
    "Verifica aggiornamenti GitHub": "Check GitHub updates",
    "Aggiornato": "Updated",
    "Errore": "Error",
    "Attenzione": "Warning",
    "Sintesi": "Summary",
    "Risultati": "Results",
    "Dettaglio": "Detail",
    "Metodo": "Method",
    "Note": "Notes",
    "Scarica": "Download",
    "Report": "Report",
    "Piano": "Plan",
    "mensile": "monthly",
    "annuale": "annual",
    "Mese": "Month",
    "Ore": "Hours",
    "Stato": "Status",
    "Validità": "Validity",
    "Validita": "Validity",
    "Conforme": "Compliant",
    "Non conforme": "Not compliant",
    "Totale": "Total",
    "Media": "Average",
    "Soglia": "Threshold",
    "Risparmio": "Saving",
    "Saving medio": "Average saving",
    "Mesi conformi": "Compliant months",
    "Produzione": "Production",
    "Energia": "Energy",
    "Ricavi": "Revenues",
    "Costi": "Costs",
    "Margine": "Margin",
    "Business Plan": "Business Plan",
    "Pro Forma": "Pro Forma",

    # Biomasse / feedstock
    "Biomassa": "Feedstock",
    "Biomasse": "Feedstocks",
    "Database feedstock": "Feedstock database",
    "Trinciato di mais": "Maize silage",
    "Trinciato di sorgo": "Sorghum silage",
    "Pollina ovaiole": "Layer poultry manure",
    "Pollina broiler": "Broiler poultry manure",
    "Liquame suino": "Pig slurry",
    "Liquame bovino": "Cattle slurry",
    "Letame bovino": "Cattle manure",
    "Letame palabile": "Solid manure",
    "FORSU": "Organic fraction of MSW",
    "Polpe barbabietola fresche": "Fresh sugar beet pulp",
    "Resa": "Yield",
    "Sottoprodotti": "By-products",
    "Effluenti": "Manure/slurry",
    "Colture dedicate": "Dedicated crops",
    "Mix biomasse": "Feedstock mix",
    "Tot. biomasse": "Total feedstocks",
    "Totale biomasse": "Total feedstocks",

    # Unità / colonne
    "Ore/mese": "Hours/month",
    "t/mese": "t/month",
    "t/anno": "t/year",
    "Sm³ lordi": "Gross Sm³",
    "Sm3 lordi": "Gross Sm3",
    "Sm³ netti": "Net Sm³",
    "Sm3 netti": "Net Sm3",
    "Sm³/h netti": "Net Sm³/h",
    "Sm3/h netti": "Net Sm3/h",
    "Sm³ CH4 lordi": "Gross CH4 Sm³",
    "Sm3 CH4 lordi": "Gross CH4 Sm3",
    "Sm³ CH4 al motore": "CH4 Sm³ to engine",
    "Sm3 CH4 al motore": "CH4 Sm3 to engine",
    "MWh netti": "Net MWh",
    "MWh el LORDI": "Gross el. MWh",
    "MWh el lordi": "Gross el. MWh",
    "MWh el netti rete": "Net grid el. MWh",
    "MWh termici": "Thermal MWh",
    "kW lordi": "Gross kW",
    "kW lordi medi": "Average gross kW",
    "GHG (gCO₂/MJ)": "GHG (gCO₂e/MJ)",
    "Saving %": "Saving %",
    "e_w": "e_w",

    # Incentivi / regimi
    "Biometano": "Biomethane",
    "Biogas": "Biogas",
    "Cogenerazione": "Cogeneration",
    "CHP": "CHP",
    "Tariffa": "Tariff",
    "Tariffa base": "Base tariff",
    "Premio": "Premium",
    "Premio matrice": "Feedstock premium",
    "Premio CAR": "HEC premium",
    "Ritiro dedicato": "Dedicated purchase",
    "Tariffa Omnicomprensiva": "All-inclusive tariff",
    "CIC": "CIC",
    "Avanzato": "Advanced",
    "Double counting": "Double counting",

    # Export
    "Excel modificabile": "Editable Excel",
    "Excel snapshot": "Excel snapshot",
    "CSV snapshot": "CSV snapshot",
    "Scarica Report PDF": "Download PDF Report",
    "Scarica risultati": "Download results",
    "Piano mensile": "Monthly plan",
    "Sintesi annuale": "Annual summary",
    "Business Plan": "Business Plan",

    # Mesi
    "Gennaio": "January",
    "Febbraio": "February",
    "Marzo": "March",
    "Aprile": "April",
    "Maggio": "May",
    "Giugno": "June",
    "Luglio": "July",
    "Agosto": "August",
    "Settembre": "September",
    "Ottobre": "October",
    "Novembre": "November",
    "Dicembre": "December",

    # Stato sintetico
    "OK": "OK",
    "KO": "KO",
    "Impossibile": "Impossible",
    "Produzione OK": "Production OK",
    "saving <": "saving <",
}

# Ordine: prima frasi lunghe, poi parole singole.
_REPLACEMENTS = sorted(IT_TO_EN.items(), key=lambda kv: len(kv[0]), reverse=True)


def translate_text(value: Any, lang: str = LANG_IT) -> Any:
    """Traduce stringhe IT->EN; lascia invariati gli altri tipi."""
    if lang != LANG_EN or not isinstance(value, str):
        return value
    if value in IT_TO_EN:
        return IT_TO_EN[value]
    out = value
    for src, dst in _REPLACEMENTS:
        out = out.replace(src, dst)
    return out


def translate_obj(value: Any, lang: str = LANG_IT) -> Any:
    """Traduce ricorsivamente strutture base usate da Streamlit."""
    if lang != LANG_EN:
        return value
    if isinstance(value, str):
        return translate_text(value, lang)
    if isinstance(value, list):
        return [translate_obj(v, lang) for v in value]
    if isinstance(value, tuple):
        return tuple(translate_obj(v, lang) for v in value)
    if isinstance(value, dict):
        return {translate_obj(k, lang): translate_obj(v, lang) for k, v in value.items()}
    return value


def translate_dataframe(df: Any, lang: str = LANG_IT) -> Any:
    """Rinomina colonne e celle testuali di un DataFrame pandas."""
    if lang != LANG_EN or pd is None:
        return df
    if not isinstance(df, pd.DataFrame):
        return df
    out = df.copy()
    out.columns = [translate_text(c, lang) for c in out.columns]
    for col in out.columns:
        if out[col].dtype == object:
            out[col] = out[col].map(lambda x: translate_text(x, lang) if isinstance(x, str) else x)
    return out


def translate_csv_bytes(data: Any, lang: str = LANG_IT) -> Any:
    """Traduce header e testi ricorrenti in CSV già serializzati."""
    if lang != LANG_EN:
        return data
    if isinstance(data, str):
        return translate_text(data, lang)
    if isinstance(data, bytes):
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                text = data.decode(enc)
                return translate_text(text, lang).encode(enc)
            except UnicodeDecodeError:
                continue
    return data


def translate_xlsx_buffer(buf: BytesIO, lang: str = LANG_IT) -> BytesIO:
    """Traduce testi visibili in un workbook XLSX generato da openpyxl."""
    if lang != LANG_EN or load_workbook is None:
        return buf
    buf.seek(0)
    wb = load_workbook(buf)
    for ws in wb.worksheets:
        ws.title = _safe_sheet_title(translate_text(ws.title, lang))
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    # Non tocchiamo formule Excel.
                    if cell.value.startswith("="):
                        continue
                    cell.value = translate_text(cell.value, lang)
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _safe_sheet_title(title: str) -> str:
    """Excel: max 31 caratteri e niente caratteri vietati."""
    bad = "[]:*?/\\"
    clean = "".join("-" if c in bad else c for c in title)
    return clean[:31] or "Sheet"


def patch_streamlit(st_module, lang_getter):
    """Applica monkeypatch leggero alle funzioni Streamlit più usate.

    Il getter consente di leggere la lingua corrente da session_state al momento
    della chiamata, senza riapplicare patch a ogni rerun.
    """
    if getattr(st_module, "_metaniq_i18n_patched", False):
        return

    def current_lang():
        try:
            return lang_getter()
        except Exception:
            return LANG_IT

    def wrap_text_func(obj, name):
        original = getattr(obj, name, None)
        if original is None or getattr(original, "_metaniq_wrapped", False):
            return

        def wrapped(*args, **kwargs):
            lang = current_lang()
            args = tuple(translate_obj(a, lang) for a in args)
            kwargs = {k: translate_obj(v, lang) for k, v in kwargs.items()}
            return original(*args, **kwargs)

        wrapped._metaniq_wrapped = True
        setattr(obj, name, wrapped)

    for target in [st_module, getattr(st_module, "sidebar", None)]:
        if target is None:
            continue
        for fn in (
            "title", "header", "subheader", "markdown", "caption", "write",
            "info", "warning", "error", "success", "button", "radio", "selectbox",
            "multiselect", "checkbox", "slider", "number_input", "text_input",
            "metric", "tabs", "expander",
        ):
            wrap_text_func(target, fn)

    # DataFrame/data_editor: prova a tradurre colonne prima del rendering.
    for name in ("dataframe", "data_editor"):
        original = getattr(st_module, name, None)
        if original is None:
            continue

        def make_df_wrapper(orig):
            def wrapped(data=None, *args, **kwargs):
                lang = current_lang()
                data = translate_dataframe(data, lang)
                kwargs = {k: translate_obj(v, lang) for k, v in kwargs.items()}
                return orig(data, *args, **kwargs)
            return wrapped

        setattr(st_module, name, make_df_wrapper(original))

    # Download button: traduce etichetta e CSV già serializzati.
    original_download = getattr(st_module, "download_button", None)
    if original_download is not None:
        def download_button(label, data=None, *args, **kwargs):
            lang = current_lang()
            label = translate_text(label, lang)
            kwargs = {k: translate_obj(v, lang) for k, v in kwargs.items()}
            mime = kwargs.get("mime", "") or ""
            if "csv" in mime.lower() or str(kwargs.get("file_name", "")).endswith(".csv"):
                data = translate_csv_bytes(data, lang)
            return original_download(label, data=data, *args, **kwargs)
        st_module.download_button = download_button

    # Column config constructors: traducono label/help.
    col_cfg = getattr(st_module, "column_config", None)
    if col_cfg is not None:
        for cls_name in (
            "TextColumn", "NumberColumn", "SelectboxColumn", "CheckboxColumn",
            "DateColumn", "DatetimeColumn", "ProgressColumn",
        ):
            original_cls = getattr(col_cfg, cls_name, None)
            if original_cls is None:
                continue

            def make_col_wrapper(orig):
                def wrapped(*args, **kwargs):
                    lang = current_lang()
                    args = tuple(translate_obj(a, lang) for a in args)
                    kwargs = {k: translate_obj(v, lang) for k, v in kwargs.items()}
                    return orig(*args, **kwargs)
                return wrapped

            setattr(col_cfg, cls_name, make_col_wrapper(original_cls))

    st_module._metaniq_i18n_patched = True
